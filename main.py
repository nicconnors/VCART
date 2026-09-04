import os
import glob
import re
import traceback
import warnings
from copy import copy
from datetime import datetime
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from config import (
    BASE_DIR, OUTPUT,
    TERRITORIES, REGION_MAP,
    SOURCE_SHEET_KEYWORDS, HEADER_ROW, DATA_START,
    SRC_DATA_START, SRC_DATA_END, NUM_DATA_COLS,
    BARRIER_MOVEMENT_SHEET_KEYWORD, TAXONOMY_SHEET_KEYWORD,
    SRC_BARRIER_BOOL_START, SRC_BARRIER_BOOL_END,
    BARRIER_TRACKING_FIELDS, DAYS_TO_CLOSE_AFTER, TERRITORY_FIELDS,
    OUT_LABEL_COL, OUT_TERRNAME_COL, OUT_TERR_COL, OUT_NAME_COL, OUT_CAMPUS_COL,
    OUT_DATA_START, MAX_OUT_COL,
    BARRIER_COLS_OUT, NATION_SUM_COLS, PCT_COLS_OUT,
    SECTION_HEADERS, TS_SECTION_HEADERS, COL_HEADERS,
    YELLOW, NATION_FILL, EAST_FILL, WEST_FILL,
    LIVE_HDR_PEACH, LIVE_HDR_GREEN, LIVE_HDR_PINK,
    VIIV_DARK_COLORS,
    MONTH_NAMES, TS_HEADER_ROW, TS_START_ROW, TS_END_ROW,
)

import summary

warnings.filterwarnings("ignore")

NUM_TERRITORIES = len(TERRITORIES)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_source_sheet(wb):
    for name in wb.sheetnames:
        nl = name.lower()
        if all(k in nl for k in SOURCE_SHEET_KEYWORDS):
            return wb[name]
    return None


def _find_barrier_movement_sheet(wb):
    """Return the first sheet whose name contains BARRIER_MOVEMENT_SHEET_KEYWORD
    (case-insensitive) — e.g. "Q3 Barrier Movement". Matched by keyword, not
    exact name, so "Q4 Barrier Movement" etc. keeps working next quarter
    without code changes, same pattern as the main sheet's own detection.
    """
    kw = BARRIER_MOVEMENT_SHEET_KEYWORD.lower()
    for name in wb.sheetnames:
        if kw in name.lower():
            return wb[name]
    return None


def _normalize_header(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def _header_col_map(ws, header_row):
    """Build {normalized_header_text: column_index} for the given header row."""
    m = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value:
                m[_normalize_header(cell.value)] = cell.column
        break
    return m


def _resolve_field_col(header_map, candidates):
    """Return the first column index matching any candidate header string, or None."""
    for cand in candidates:
        key = _normalize_header(cand)
        if key in header_map:
            return header_map[key]
    return None


# Excel formula-error artifacts — a source cell holding one of these (e.g.
# #DIV/0! from a Days-to-Close AVERAGE formula dividing by zero closed
# barriers) gets read as this literal string by openpyxl; left unhandled
# it flows straight through into the output and shows up verbatim in a
# client-facing report. The old code only ever caught "#N/A" specifically
# — ported from ADFR's fuller set.
EXCEL_ERROR_STRINGS = {"#n/a", "#div/0!", "#ref!", "#value!", "#name?", "#null!", "#num!"}


def clean_source_value(val):
    """Return None in place of an Excel formula-error string; pass everything else through."""
    if isinstance(val, str) and val.strip().lower() in EXCEL_ERROR_STRINGS:
        return None
    return val


def compute_total_barriers(all_td):
    """Nationwide total barrier count — sum of BARRIER_COLS_OUT across every
    loaded territory. Used everywhere a 'total barriers' figure is shown
    (LIVE label, TS row, final console summary) so all four previously used
    slightly different filters here — this is the one place that formula
    lives now, so they can't silently disagree again. bool is explicitly
    excluded even though isinstance(True, int) is True in Python, in case a
    source cell ever holds a literal TRUE/FALSE instead of a count.
    """
    return sum(
        sum(td["totals"].get(c, 0) or 0 for c in BARRIER_COLS_OUT
            if isinstance(td["totals"].get(c), (int, float)) and not isinstance(td["totals"].get(c), bool))
        for td in all_td.values() if td
    )


def _section_fill(out_col, headers=None):
    if headers is None:
        headers = SECTION_HEADERS
    for start, end, _, color in headers:
        if start <= out_col <= end:
            return PatternFill("solid", fgColor=color)
    return PatternFill(fill_type=None)


def _shift_section_headers(headers, offset):
    """
    SECTION_HEADERS column ranges are calibrated for the VCART Totals sheet,
    where data columns start at OUT_DATA_START (6). The VCART Systems sheets
    have 8 demographic columns (not 5), so their data starts 3 columns later
    at SRC_DATA_START (9). Passing SECTION_HEADERS to write_section_headers()
    unshifted paints the row-3 color band 3 columns too far left — bleeding
    color onto City/State/Zip and cutting off the last 3 columns of each
    section. This shifts every range by `offset` so the band lines up with
    the actual columns used on that sheet.
    """
    return [(start + offset, end + offset, label, color) for start, end, label, color in headers]


# Systems sheets (VCART Systems - LIVE / snapshot) have their data columns
# shifted right by this many columns relative to the VCART Totals sheet.
SYSTEMS_COL_OFFSET = SRC_DATA_START - OUT_DATA_START  # 9 - 6 = 3
SYSTEMS_SECTION_HEADERS = _shift_section_headers(SECTION_HEADERS, SYSTEMS_COL_OFFSET)

# The top time-series table's own "Months" / "Total Barriers" label columns.
# Per request, these sit at D/E now (cols 1-3 left blank as a spacer above
# the LIVE/region table, which still uses cols 1-5 for Region/Territory
# Name/Territory #/VCART Name/Total Accounts further down the same sheet).
# Purely a cosmetic repositioning of the same two labels — data columns
# still start at OUT_DATA_START and are untouched.
TS_MONTH_COL     = OUT_DATA_START - 2   # 4 (D)
TS_BARRIERS_COL  = OUT_DATA_START - 1   # 5 (E)

# Barrier 1/2 Start Date and Close Date columns — hidden per request. Days
# to Close (16/21 = P/U) stays visible; only the raw date columns hide.
HIDDEN_BARRIER_DATE_COLS = (14, 15, 19, 20)  # N, O, S, T

# Same fields, shifted for "VCART Systems - LIVE"'s wider demographic
# block (8 cols instead of 5) — Q, R, V, W. Days to Close (S/X) stays
# visible there too.
HIDDEN_SYSTEMS_BARRIER_DATE_COLS = tuple(c + SYSTEMS_COL_OFFSET for c in HIDDEN_BARRIER_DATE_COLS)

# Barrier 1/2 Days to Close output columns on VCART Totals.
BARRIER_DAYS_TO_CLOSE_COLS = (16, 21)

# All-borders styling used by finalize_workbook_formatting()
ALL_BORDER_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(left=ALL_BORDER_SIDE, right=ALL_BORDER_SIDE,
                      top=ALL_BORDER_SIDE, bottom=ALL_BORDER_SIDE)


# ---------------------------------------------------------------------------
# FILE READING
# ---------------------------------------------------------------------------

def _read_row3(ws):
    """Return row 3 as a plain list of values (positional, 0-based)."""
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        return list(row)
    return []


def _row_values(ws, row_num):
    """Return a 1-indexed {col: value} dict for one row."""
    rows = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    if not rows:
        return {}
    return {i + 1: v for i, v in enumerate(rows[0])}


def _campus_identity(row):
    """CID (col C, 0-indexed 2) if it's a real int, else None. Used to
    match a campus between the main sheet and the Barrier Movement sheet.
    """
    cid = row[2] if len(row) > 2 else None
    return cid if isinstance(cid, int) else None


def _campus_name_key(row):
    """Normalized Corporate Parent Name (col D, 0-indexed 3), used as a
    fallback identity when CID isn't populated at all — some territory
    files track campuses by name instead of CID (e.g. 'DAP', 'LGBT',
    'AltaMed'), and in those files CID is blank on every row of BOTH
    sheets, so CID matching can never work for that territory regardless
    of how correct the Barrier Movement sheet is.
    """
    name = row[3] if len(row) > 3 else None
    if not name:
        return None
    return re.sub(r"\s+", " ", str(name)).strip().lower()


def _is_real_campus_row(row, require_bool_col=True):
    """
    A real campus data row (not a blank template row) needs a real
    identity (Territory #, VCART Name, or a genuine CID integer).

    require_bool_col additionally checks that col I (index 8) looks like a
    real barrier-boolean answer — valid on the Barrier Movement sheet
    (col I is genuinely "Access for All" there), but NOT on the main
    sheet, where col I is "Centralized or Decentralized..." (a string)
    since barrier data moved off that sheet. Using this check
    unconditionally on main-sheet rows made every single row fail it —
    campus_rows came back empty for every territory, silently, since the
    barrier-sheet split happened.
    """
    has_identity = (
        (row[0] and str(row[0]).strip()) or
        (row[1] and str(row[1]).strip()) or
        (row[2] and isinstance(row[2], int))
    )
    if not has_identity:
        return False
    if not require_bool_col:
        return True
    # col I (index 8) must look like a real barrier-boolean answer
    return len(row) > 8 and isinstance(row[8], (bool, int, float))


# =============================================================================
# QUARTER MAPPING
# Calendar-quarter helper used by the "closed this quarter" scope below —
# Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec. Kept as its own function
# (rather than inlined) so the quarter definition lives in exactly one
# place if it ever needs to change (e.g. a fiscal year offset). Ported
# from ADFR's identical helper.
# =============================================================================

def get_quarter_months(month):
    """Return the set of calendar months (1-12) in the quarter containing `month`.
    Q1=Jan-Mar, Q2=Apr-Jun, Q3=Jul-Sep, Q4=Oct-Dec."""
    start = ((month - 1) // 3) * 3 + 1
    return set(range(start, start + 3))


def scan_campus_barrier_closures(ws_mv, mv_header_map, target_year=None, target_month=None,
                                  scope="month"):
    """
    Scan per-campus rows on the Barrier Movement sheet. A barrier is
    "closed" if its Close Date cell holds a real date — text entries
    don't count. Ported from ADFR's version of the same function.

    scope controls the date window used to decide which closures count:
      "month"   — only counts a barrier closed in target_month/target_year
                  (defaults to the current real month/year). This is
                  "closed this month," and re-runs later in the same
                  calendar month will always agree with each other.
      "quarter" — only counts a barrier closed within the same calendar
                  quarter as target_month (Q1=Jan-Mar, Q2=Apr-Jun,
                  Q3=Jul-Sep, Q4=Oct-Dec) AND target_year (defaults to the
                  current real month/year, so "this quarter" means the
                  quarter containing today). A close date from a different
                  quarter — even one still sitting in this quarter's
                  source file, e.g. a stale row that never got cleaned up
                  — is correctly excluded. This does NOT assume "every
                  closed barrier in the file belongs to the current
                  quarter"; it checks the actual Close Date against the
                  quarter's month range.
      "all"     — every closed barrier counts, no date filtering at all.

    Returns (closed_count, category_days, barrier1_days, barrier2_days):
      closed_count  — barriers closed (per scope above), this territory
                       (Barrier 1 + 2)
      category_days — list of (category_name, days_to_close) for every
                       counted barrier that also has a resolved category
                       and a valid Days to Close number, for both slots
      barrier1_days — plain list of Days to Close values, Barrier 1 slot
                       only, for every counted barrier with a valid number
      barrier2_days — same, Barrier 2 slot only
    """
    if scope in ("month", "quarter") and (target_year is None or target_month is None):
        now = datetime.now()
        target_year = target_year if target_year is not None else now.year
        target_month = target_month if target_month is not None else now.month

    allowed_months = get_quarter_months(target_month) if scope == "quarter" else None

    cat1_col   = _resolve_field_col(mv_header_map, BARRIER_TRACKING_FIELDS[12])
    close1_col = _resolve_field_col(mv_header_map, BARRIER_TRACKING_FIELDS[15])
    days1_col  = close1_col + 1 if close1_col else None
    cat2_col   = _resolve_field_col(mv_header_map, BARRIER_TRACKING_FIELDS[17])
    close2_col = _resolve_field_col(mv_header_map, BARRIER_TRACKING_FIELDS[20])
    days2_col  = close2_col + 1 if close2_col else None

    closed_count = 0
    category_days = []
    barrier1_days = []
    barrier2_days = []

    def _at(row_vals, col):
        if col is None or col - 1 >= len(row_vals):
            return None
        return row_vals[col - 1]

    for row in ws_mv.iter_rows(min_row=DATA_START, values_only=True):
        if not _is_real_campus_row(row):
            continue
        for slot_idx, (close_col, cat_col, days_col) in enumerate((
                (close1_col, cat1_col, days1_col),
                (close2_col, cat2_col, days2_col))):
            close_val = _at(row, close_col)
            if not isinstance(close_val, datetime):
                continue
            if scope == "month":
                if close_val.year != target_year or close_val.month != target_month:
                    continue
            elif scope == "quarter":
                if close_val.year != target_year or close_val.month not in allowed_months:
                    continue
            # scope == "all": no date filtering at all.
            closed_count += 1
            cat_val = _at(row, cat_col)
            days_val = clean_source_value(_at(row, days_col))
            # A negative days-to-close means the source's Close Date was
            # entered before its Start Date — a data-entry error, not a
            # real duration. Still counts as closed, but excluded from
            # the category-average pairing and the barrier1/2 lists below.
            if isinstance(days_val, (int, float)) and days_val >= 0:
                if slot_idx == 0:
                    barrier1_days.append(days_val)
                else:
                    barrier2_days.append(days_val)
                if isinstance(cat_val, str) and cat_val.strip():
                    category_days.append((cat_val.strip(), days_val))

    return closed_count, category_days, barrier1_days, barrier2_days


def read_territory_file(filepath, terr_label):
    """
    Read one territory VCART file.
    Returns:
        terr_num    - territory number string
        vcart_name  - VCART rep name
        campus_rows - list of dicts {out_col: value} for each data campus row
        totals_row  - dict {out_col: value} from the pre-computed TOTALS rows
    """
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening {os.path.basename(filepath)}: {e}")
        return None, None, [], {}

    ws_main = _find_source_sheet(wb)
    if ws_main is None:
        print(f"  WARNING: no VRHR.VCART sheet found in {os.path.basename(filepath)}")
        wb.close()
        return None, None, [], {}

    ws_mv = _find_barrier_movement_sheet(wb)
    if ws_mv is None:
        print(f"  WARNING: No '{BARRIER_MOVEMENT_SHEET_KEYWORD}' sheet found in "
              f"{os.path.basename(filepath)} — barrier booleans and tracking "
              f"fields (cols {OUT_DATA_START}-21) will be blank for this territory")

    # ------------------------------------------------------------------
    # Read TOTALS rows (row 3)
    # ------------------------------------------------------------------
    main_row3 = _read_row3(ws_main)
    totals_raw = {}

    # Territory fields (Centralized/Decentralized onward): header-matched, main sheet — see
    # config.py's TERRITORY_FIELDS note on why this can't be positional
    # (two real territory files in the same quarter don't even agree on
    # where these columns sit).
    main_header_map = _header_col_map(ws_main, HEADER_ROW)
    # Resolved once per file (not per row) — the header map doesn't change
    # within a file, so re-running _resolve_field_col for every campus row
    # was repeating the same lookup unnecessarily on territories with many
    # campuses. Reused below both for the TOTALS row and every campus row.
    territory_resolved_src_cols = {
        out_col: _resolve_field_col(main_header_map, candidates)
        for out_col, candidates in TERRITORY_FIELDS.items()
    }
    for out_col, candidates in TERRITORY_FIELDS.items():
        src_col = territory_resolved_src_cols[out_col]
        if src_col is None:
            print(f"  WARNING: Could not find header for output col {out_col} "
                  f"(tried: {candidates}) on main sheet in "
                  f"{os.path.basename(filepath)} — leaving blank")
            totals_raw[out_col] = None
        else:
            v = main_row3[src_col - 1] if src_col - 1 < len(main_row3) else None
            totals_raw[out_col] = clean_source_value(v)

    mv_header_map = {}
    mv_resolved_src_cols = {}
    if ws_mv is not None:
        mv_row3 = _read_row3(ws_mv)
        mv_header_map = _header_col_map(ws_mv, HEADER_ROW)
        # Resolved once per file, same reasoning as territory_resolved_src_cols
        # above — reused for the TOTALS row and every campus row below.
        mv_resolved_src_cols = {
            out_col: _resolve_field_col(mv_header_map, candidates)
            for out_col, candidates in BARRIER_TRACKING_FIELDS.items()
        }

        # Barrier booleans: positional, off the Barrier Movement sheet.
        for src_idx in range(SRC_BARRIER_BOOL_START - 1, SRC_BARRIER_BOOL_END):  # 0-based
            v = mv_row3[src_idx] if src_idx < len(mv_row3) else None
            out_col = OUT_DATA_START + (src_idx - (SRC_BARRIER_BOOL_START - 1))
            totals_raw[out_col] = clean_source_value(v)

        # Barrier Tracking fields: header-matched, off the Barrier Movement sheet.
        for out_col, candidates in BARRIER_TRACKING_FIELDS.items():
            src_col = mv_resolved_src_cols[out_col]
            if src_col is None:
                print(f"  WARNING: Could not find header for output col {out_col} "
                      f"(tried: {candidates}) on Barrier Movement sheet in "
                      f"{os.path.basename(filepath)} — leaving blank")
                totals_raw[out_col] = None
            else:
                v = mv_row3[src_col - 1] if src_col - 1 < len(mv_row3) else None
                totals_raw[out_col] = clean_source_value(v)

        # Days to close: one column to the right of its paired Close Date
        # column (matched by header text just above), not by header text
        # itself, since "Days to close" isn't unique across Barrier 1 / 2.
        for dtc_out_col, close_date_out_col in DAYS_TO_CLOSE_AFTER.items():
            close_src_col = mv_resolved_src_cols.get(close_date_out_col)
            if close_src_col is None:
                totals_raw[dtc_out_col] = None
                continue
            dtc_src_col = close_src_col + 1
            v = mv_row3[dtc_src_col - 1] if dtc_src_col - 1 < len(mv_row3) else None
            totals_raw[dtc_out_col] = clean_source_value(v)
    else:
        for out_col in range(OUT_DATA_START, 24):
            totals_raw[out_col] = None

    # ------------------------------------------------------------------
    # Read data rows (row 7 onward) from the MAIN sheet — demographics +
    # CEP-onward fields live here, and this remains the authoritative
    # source of which campuses exist.
    # ------------------------------------------------------------------
    vcart_name = None
    terr_num   = None
    campus_rows = []

    main_rows = list(ws_main.iter_rows(min_row=DATA_START, values_only=True))

    # Build a CID -> row lookup for the Barrier Movement sheet, so campus
    # rows are matched by identity rather than assumed to be in the same
    # row order on both sheets (they happen to match in practice, but
    # matching by CID is just as cheap and doesn't silently break if a
    # future edit reorders one sheet but not the other).
    mv_by_cid = {}
    mv_by_name = {}
    name_seen_count = Counter()
    if ws_mv is not None:
        for row in ws_mv.iter_rows(min_row=DATA_START, values_only=True):
            if all(v is None or v == "" for v in row[:SRC_BARRIER_BOOL_END]):
                continue
            if not _is_real_campus_row(row):
                continue
            cid = _campus_identity(row)
            if cid is not None:
                mv_by_cid[cid] = row
            name_key = _campus_name_key(row)
            if name_key is not None:
                name_seen_count[name_key] += 1
                mv_by_name[name_key] = row

    unmatched_count = 0
    ambiguous_name_count = 0

    for row in main_rows:
        # Stop at fully empty row
        if all(v is None or v == "" for v in row[:SRC_DATA_END]):
            continue

        # Territory # from col A (0-indexed: 0)
        if terr_num is None:
            if row[0] and str(row[0]).strip():
                terr_num = str(row[0]).strip()

        # VCART name from col B (index 1)
        if vcart_name is None and row[1] and str(row[1]).strip():
            vcart_name = str(row[1]).strip()

        if not _is_real_campus_row(row, require_bool_col=False):
            continue  # blank template rows — skip

        rd = {}
        # Store demographics: cols A-H (indices 0-7)
        rd["terr_num"]   = row[0] if row[0] else terr_num
        rd["vcart_name"] = row[1] if row[1] else vcart_name
        rd["cid"]        = row[2] if len(row) > 2 else None
        rd["corp_name"]  = row[3] if len(row) > 3 else None
        rd["address"]    = row[4] if len(row) > 4 else None
        rd["city"]       = row[5] if len(row) > 5 else None
        rd["state"]      = row[6] if len(row) > 6 else None
        rd["zip"]        = row[7] if len(row) > 7 else None

        # Territory fields (Centralized/Decentralized onward): header-matched, main sheet —
        # reuses territory_resolved_src_cols resolved once above the row loop.
        for out_col, candidates in TERRITORY_FIELDS.items():
            src_col = territory_resolved_src_cols[out_col]
            v = row[src_col - 1] if (src_col and src_col - 1 < len(row)) else None
            rd[out_col] = clean_source_value(v)

        # Barrier booleans + tracking: from the matching Barrier Movement
        # campus row. Matched by CID when available; falls back to
        # Corporate Parent Name when CID is blank on both sheets (some
        # territory files identify campuses by name instead of CID).
        cid = _campus_identity(row)
        mv_row = mv_by_cid.get(cid) if (ws_mv is not None and cid is not None) else None
        matched_by_name = False
        if mv_row is None and ws_mv is not None:
            name_key = _campus_name_key(row)
            if name_key is not None:
                if name_seen_count.get(name_key, 0) == 1:
                    mv_row = mv_by_name.get(name_key)
                    matched_by_name = mv_row is not None
                elif name_seen_count.get(name_key, 0) > 1:
                    ambiguous_name_count += 1
        if mv_row is None:
            if ws_mv is not None:
                unmatched_count += 1
            for out_col in range(OUT_DATA_START, 24):
                rd[out_col] = None
        else:
            for src_idx in range(SRC_BARRIER_BOOL_START - 1, SRC_BARRIER_BOOL_END):
                v = mv_row[src_idx] if src_idx < len(mv_row) else None
                out_col = OUT_DATA_START + (src_idx - (SRC_BARRIER_BOOL_START - 1))
                rd[out_col] = clean_source_value(v)
            # reuses mv_resolved_src_cols resolved once above the row loop.
            for out_col, candidates in BARRIER_TRACKING_FIELDS.items():
                src_col = mv_resolved_src_cols.get(out_col)
                v = mv_row[src_col - 1] if (src_col and src_col - 1 < len(mv_row)) else None
                rd[out_col] = clean_source_value(v)
            for dtc_out_col, close_date_out_col in DAYS_TO_CLOSE_AFTER.items():
                close_src_col = mv_resolved_src_cols.get(close_date_out_col)
                if close_src_col is None:
                    rd[dtc_out_col] = None
                    continue
                dtc_src_col = close_src_col + 1
                v = mv_row[dtc_src_col - 1] if dtc_src_col - 1 < len(mv_row) else None
                rd[dtc_out_col] = clean_source_value(v)

        campus_rows.append(rd)

    if unmatched_count:
        print(f"  WARNING: {unmatched_count} campus row(s) in {os.path.basename(filepath)} "
              f"had no matching CID or Corporate Parent Name on the Barrier Movement sheet "
              f"— barrier fields left blank for those campuses")
    if ambiguous_name_count:
        print(f"  WARNING: {ambiguous_name_count} campus row(s) in {os.path.basename(filepath)} "
              f"had a Corporate Parent Name that appears more than once on the Barrier "
              f"Movement sheet (and no CID to disambiguate) — barrier fields left blank "
              f"rather than risk matching the wrong campus")

    if terr_num is None:
        terr_num = "UNKNOWN"
    if vcart_name is None:
        vcart_name = terr_label

    # Closed-barrier tracking — per-campus scan, not the TOTALS row (which
    # has no closed/open concept and, per the source file, can even hold a
    # #DIV/0! formula-error artifact for Days to Close).
    if ws_mv is not None:
        closed_count, category_days, barrier1_days, barrier2_days = scan_campus_barrier_closures(
            ws_mv, mv_header_map, scope="month")
        totals_raw["closed_count"] = closed_count
        totals_raw["category_days_to_close"] = category_days
        totals_raw["barrier1_days_to_close"] = barrier1_days
        totals_raw["barrier2_days_to_close"] = barrier2_days

        # Scoped to the current calendar quarter (Jan-Mar/Apr-Jun/Jul-Sep/
        # Oct-Dec) — a barrier only counts here if its own Close Date
        # actually falls within that 3-month window AND the current year,
        # not just because it's a closed barrier that happens to be
        # sitting in this quarter's source file. A stale close date from
        # an earlier quarter (or year) is excluded.
        (closed_count_quarter, category_days_quarter,
         barrier1_days_quarter, barrier2_days_quarter) = scan_campus_barrier_closures(
            ws_mv, mv_header_map, scope="quarter")
        totals_raw["closed_count_quarter"] = closed_count_quarter
        totals_raw["category_days_to_close_quarter"] = category_days_quarter
        totals_raw["barrier1_days_to_close_quarter"] = barrier1_days_quarter
        totals_raw["barrier2_days_to_close_quarter"] = barrier2_days_quarter
    else:
        totals_raw["closed_count"] = 0
        totals_raw["category_days_to_close"] = []
        totals_raw["barrier1_days_to_close"] = []
        totals_raw["barrier2_days_to_close"] = []
        totals_raw["closed_count_quarter"] = 0
        totals_raw["category_days_to_close_quarter"] = []
        totals_raw["barrier1_days_to_close_quarter"] = []
        totals_raw["barrier2_days_to_close_quarter"] = []

    wb.close()
    return terr_num, vcart_name, campus_rows, totals_raw


def load_barrier_taxonomy(folder):
    """
    Return {normalized_subcategory_text: category_display_name}.
    Reads the "Lists" sheet from the first territory file that loads
    successfully: each of the first 6 columns is headed with a category
    name (row 1) and lists that category's subcategories below it — same
    structure as ADFR's version, confirmed against real VCART files.
    Falls back to an empty dict (subcategories show as "Uncategorized" in
    the % table) if no source file with a Lists sheet can be found.
    """
    for _, filename in TERRITORIES:
        fp = os.path.join(folder, filename)
        if not os.path.exists(fp):
            matches = glob.glob(os.path.join(folder, f"*{filename.split('_')[0]}*"))
            if not matches:
                continue
            fp = matches[0]
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
        except Exception:
            continue
        ws = None
        for name in wb.sheetnames:
            if TAXONOMY_SHEET_KEYWORD in name.lower():
                ws = wb[name]
                break
        if ws is None:
            wb.close()
            continue

        taxonomy = {}
        header_row = _row_values(ws, 1)
        for col in range(1, 7):  # first 6 columns = the 6 barrier categories
            category = header_row.get(col)
            if not category:
                continue
            category = str(category).strip()
            for r in range(2, ws.max_row + 1):
                sub = ws.cell(row=r, column=col).value
                if sub and str(sub).strip():
                    taxonomy[_normalize_header(sub)] = category
        wb.close()
        if taxonomy:
            print(f"  Loaded barrier taxonomy from {os.path.basename(fp)} "
                  f"({len(taxonomy)} subcategories across 6 categories)")
            return taxonomy

    print("  WARNING: Could not load barrier taxonomy from any source file — "
          "subcategories will show as 'Uncategorized' in the % breakdown table")
    return {}


def load_all_territories(folder):
    """
    Load all territory files. Returns dict:
        label → {terr_num, vcart_name, campus_rows, totals, campus_count, region}
    """
    result = {}
    for label, filename in TERRITORIES:
        filepath = os.path.join(folder, filename)
        if not os.path.exists(filepath):
            matches = glob.glob(os.path.join(folder, f"*{filename.split('_')[0]}*"))
            if matches:
                filepath = matches[0]
                print(f"  NOTE: fuzzy matched {label} → {os.path.basename(filepath)}")
            else:
                print(f"  WARNING: file not found for {label}: {filename}")
                result[label] = None
                continue

        print(f"  Reading {label}...")
        terr_num, vcart_name, campus_rows, totals = read_territory_file(
            filepath, label
        )
        if campus_rows is None and not totals:
            result[label] = None
            continue

        # Use actual campus count from data rows
        campus_count = len(campus_rows)
        result[label] = {
            "terr_num":    terr_num,
            "vcart_name":  vcart_name,
            "campus_rows": campus_rows,
            "totals":      totals,
            "campus_count": campus_count,
            "region":      REGION_MAP.get(label, "VCART - Unknown"),
        }
        print(f"    ✓ {label}: {campus_count} campuses (terr: {terr_num})")
    return result


# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

def unmerge_row(ws, row):
    to_remove = [str(m) for m in ws.merged_cells.ranges
                 if m.min_row <= row <= m.max_row]
    for m in to_remove:
        ws.unmerge_cells(m)


def get_last_ts_row(ws):
    last = TS_HEADER_ROW
    for r in range(TS_START_ROW, TS_END_ROW + 1):
        val = ws.cell(row=r, column=TS_MONTH_COL).value
        if val and any(m.lower() in str(val).lower() for m in MONTH_NAMES):
            last = r
    return last


def get_live_base(ws):
    return get_last_ts_row(ws) + 2


def get_live_offsets(ws):
    base = get_live_base(ws)
    raw_nat = base + 3 + NUM_TERRITORIES
    return {
        "live_row":      base,
        "raw_sec_row":   base + 1,
        "raw_hdr_row":   base + 2,
        "raw_data_row":  base + 3,
        "raw_nat_row":   raw_nat,
        "gap_row":       raw_nat + 1,
        "pct_label_row": raw_nat + 2,
        "pct_sec_row":   raw_nat + 3,
        "pct_hdr_row":   raw_nat + 4,
        "pct_data_row":  raw_nat + 5,
        "pct_nat_row":   raw_nat + 5 + NUM_TERRITORIES,
    }


def get_snap_start(ws):
    o = get_live_offsets(ws)
    return o["pct_nat_row"] + 4


def is_yellow_row(ws, row):
    try:
        fill = ws.cell(row=row, column=1).fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            return rgb in ("FFFFFF00", "00FFFF00", "FFFF00", f"FF{YELLOW}", YELLOW)
    except Exception:
        pass
    return False


def find_yellow_rows(ws, start_row, end_row):
    yellow_rows = []
    for r in range(start_row, end_row + 1):
        if is_yellow_row(ws, r):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip():
                yellow_rows.append(r)
    return yellow_rows


def _autofit_column_widths(ws, min_width=6, max_width=32, padding=2):
    """Set each column's width from the content actually in it. See VCART
    Systems sheet build function for full rationale.
    """
    merged_coords = set()
    for mr in ws.merged_cells.ranges:
        for row in ws.iter_rows(min_row=mr.min_row, max_row=mr.max_row,
                                 min_col=mr.min_col, max_col=mr.max_col):
            for cell in row:
                merged_coords.add(cell.coordinate)

    data_widths = {}
    header_word_floor = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.coordinate in merged_coords:
                continue
            text = str(cell.value)
            is_header = bool(cell.alignment and cell.alignment.wrap_text)
            if is_header:
                words = re.split(r"[\s\n]+", text)
                longest_word = max((len(w) for w in words if w), default=0)
                header_word_floor[cell.column] = max(header_word_floor.get(cell.column, 0), longest_word)
            else:
                longest_line = max((len(line) for line in text.split("\n")), default=0)
                data_widths[cell.column] = max(data_widths.get(cell.column, 0), longest_line)

    for col in set(data_widths) | set(header_word_floor):
        length = max(data_widths.get(col, 0), header_word_floor.get(col, 0))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(length + padding, max_width))


# ---------------------------------------------------------------------------
# TABLE WRITERS — VCART Totals sheet
# ---------------------------------------------------------------------------

def write_section_headers(ws, row, headers=None):
    if headers is None:
        headers = SECTION_HEADERS
    unmerge_row(ws, row)
    for start, end, label, color in headers:
        text_color = "FFFFFF" if color in VIIV_DARK_COLORS else "000000"
        cell = ws.cell(row=row, column=start, value=label)
        cell.font = Font(bold=True, size=9, color=text_color)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(start + 1, end + 1):
            ws.cell(row=row, column=c).value = None
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=color)
        if end > start:
            try:
                ws.merge_cells(start_row=row, start_column=start,
                               end_row=row, end_column=end)
            except Exception:
                pass
    ws.row_dimensions[row].height = 20


def write_live_col_headers(ws, row):
    """Write LIVE section column headers with custom per-zone colors."""
    for col, text in COL_HEADERS.items():
        cell = ws.cell(row=row, column=col, value=text)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if 1 <= col <= 5:
            fill_color = LIVE_HDR_PEACH
        elif 6 <= col <= 11:
            fill_color = LIVE_HDR_GREEN
        elif 12 <= col <= 23:
            fill_color = LIVE_HDR_PINK
        else:
            fill_color = None
        if fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)
        else:
            cell.fill = _section_fill(col)
            fill_color = next((c for s, e, _, c in SECTION_HEADERS if s <= col <= e), None)
        text_color = "FFFFFF" if fill_color in VIIV_DARK_COLORS else "000000"
        cell.font = Font(size=9, color=text_color)
    ws.row_dimensions[row].height = 80


def write_col_headers(ws, row, headers=None):
    if headers is None:
        headers = SECTION_HEADERS
    for col, text in COL_HEADERS.items():
        cell = ws.cell(row=row, column=col, value=text)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = _section_fill(col, headers)
        fill_color = next((c for s, e, _, c in headers if s <= col <= e), None)
        text_color = "FFFFFF" if fill_color in VIIV_DARK_COLORS else "000000"
        cell.font = Font(size=9, color=text_color)
    ws.row_dimensions[row].height = 80


def write_territory_row(ws, row_num, label, td, raw=True):
    """Write one territory row to the VCART Totals sheet."""
    for c in range(1, MAX_OUT_COL + 1):
        ws.cell(row=row_num, column=c).value = None
        ws.cell(row=row_num, column=c).fill = PatternFill(fill_type=None)

    if td is None:
        ws.cell(row=row_num, column=OUT_LABEL_COL, value=label)
        return

    region = td["region"]
    rfill  = PatternFill(fill_type=None)  # white — all data rows

    def _set(col, val, bold=False, fmt=None):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(size=11, bold=bold)
        is_number = isinstance(val, (int, float))
        cell.alignment = Alignment(horizontal="center" if is_number else "left",
                                    vertical="center", wrap_text=False)
        cell.fill = rfill
        if fmt:
            cell.number_format = fmt

    _set(OUT_LABEL_COL,   td["region"])
    _set(OUT_TERRNAME_COL, label)
    _set(OUT_TERR_COL,    td["terr_num"])
    _set(OUT_NAME_COL,    td["vcart_name"])
    _set(OUT_CAMPUS_COL,  td["campus_count"])

    campus_count = td["campus_count"]
    totals = td["totals"]  # pre-computed sums from file's TOTALS row

    for out_col in range(OUT_DATA_START, MAX_OUT_COL + 1):
        val = totals.get(out_col)
        cell = ws.cell(row=row_num, column=out_col)
        cell.font = Font(size=11)
        cell.fill = rfill
        cell.number_format = "General"  # explicit every time — never inherit a stale "0%" from a prior run

        if val is None or str(val) in ("#N/A", "None"):
            cell.value = None
        elif not raw and out_col in BARRIER_COLS_OUT and campus_count and isinstance(val, (int, float)):
            cell.value = val / campus_count
            cell.number_format = "0%"
        elif not raw and out_col in PCT_COLS_OUT and isinstance(val, (int, float)):
            cell.value = val  # already a pct (0–1) or store as-is
            cell.number_format = "0%"
        else:
            cell.value = val

        # Numbers center, text left — decided after the final value/type
        # above is known, since raw vs. pct mode changes what ends up here.
        is_number = isinstance(cell.value, (int, float))
        cell.alignment = Alignment(horizontal="center" if is_number else "left",
                                    vertical="center", wrap_text=False)

    ws.row_dimensions[row_num].height = 15


def write_nation_row(ws, row_num, all_td, raw=True):
    """Write NATION summary row."""
    for c in range(1, MAX_OUT_COL + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.value = None
        cell.number_format = "General"  # explicit every time — never inherit a stale "0%" from a prior run

    nation_fill = PatternFill("solid", fgColor=NATION_FILL)

    def _set(col, val, bold=True, fmt=None):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(size=11, bold=bold)
        cell.fill = nation_fill
        is_number = isinstance(val, (int, float))
        cell.alignment = Alignment(horizontal="center" if is_number else "left", vertical="center")
        if fmt:
            cell.number_format = fmt

    total_campuses = sum(td["campus_count"] for td in all_td.values() if td)
    _set(OUT_LABEL_COL, "NATION")
    _set(OUT_TERRNAME_COL, "")
    _set(OUT_TERR_COL,  "")
    _set(OUT_NAME_COL,  "")
    _set(OUT_CAMPUS_COL, total_campuses)

    valid = [td for td in all_td.values() if td]
    for out_col in NATION_SUM_COLS:
        vals = [td["totals"].get(out_col) for td in valid
                if td["totals"].get(out_col) is not None]
        numeric = [v for v in vals if isinstance(v, (int, float))]
        cell = ws.cell(row=row_num, column=out_col)
        cell.font = Font(size=11, bold=True)
        cell.fill = nation_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if numeric:
            total = sum(numeric)
            if raw:
                cell.value = total
            elif total_campuses:
                cell.value = total / total_campuses
                cell.number_format = "0%"

    # PCT_COLS_OUT fields (e.g. "% campuses flagged for TFRM") are already
    # expressed as a per-territory percentage, not a raw count — summing
    # them the way NATION_SUM_COLS does would produce a meaningless number
    # (a sum of percentages). NATION's value here is the campus-weighted
    # average across territories: sum(pct_i * campus_count_i) / total
    # campuses, which correctly reduces to the same % if every territory
    # reports the same rate, and is well-defined when they don't.
    for out_col in PCT_COLS_OUT:
        weighted_sum = 0.0
        weight_total = 0
        for td in valid:
            v = td["totals"].get(out_col)
            if isinstance(v, (int, float)) and td["campus_count"]:
                weighted_sum += v * td["campus_count"]
                weight_total += td["campus_count"]
        if weight_total:
            cell = ws.cell(row=row_num, column=out_col)
            cell.font = Font(size=11, bold=True)
            cell.fill = nation_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = weighted_sum / weight_total
            cell.number_format = "0%"

    # Days to Close (16, 21) — pooled average of every individual closed
    # barrier's own Days to Close number, nationwide, not an average of
    # each territory's own pre-existing value. Barrier 1 and Barrier 2 are
    # kept separate (col 16 only pools Barrier 1 closures, col 21 only
    # Barrier 2). Uses the "closed this quarter" lists — properly windowed
    # to the current calendar quarter (see scan_campus_barrier_closures) —
    # matching the scope each territory's own value used to reflect. Same
    # pooled average shown in both raw and pct NATION rows — not a "% of
    # total" field, just a plain number.
    b1_pool = [d for td in valid for d in td["totals"].get("barrier1_days_to_close_quarter", [])]
    b2_pool = [d for td in valid for d in td["totals"].get("barrier2_days_to_close_quarter", [])]
    for out_col, pool in zip(BARRIER_DAYS_TO_CLOSE_COLS, (b1_pool, b2_pool)):
        if pool:
            cell = ws.cell(row=row_num, column=out_col)
            cell.font = Font(size=11, bold=True)
            cell.fill = nation_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = sum(pool) / len(pool)
            cell.number_format = "0"

    ws.row_dimensions[row_num].height = 15


# ---------------------------------------------------------------------------
# VCART Totals sheet — structure
# ---------------------------------------------------------------------------

def write_totals_sheet_headers(ws):
    # Title sits over the TS table's own D:E columns (TS_MONTH_COL /
    # TS_BARRIERS_COL) — cols A-C stay blank above the LIVE/region table
    # that uses them further down this sheet. Unmerge any legacy A1:B2
    # title merge from before this layout existed, so re-running on an
    # older file corrects it instead of leaving a stale merge behind.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= 2 and mr.max_row <= 2 and mr.min_col == 1:
            ws.unmerge_cells(str(mr))
    for c in range(1, TS_MONTH_COL):
        ws.cell(row=1, column=c).value = None
    ws.cell(row=1, column=TS_MONTH_COL, value="VCART Territory Totals").font = Font(bold=True, size=14)
    try:
        ws.merge_cells(start_row=1, start_column=TS_MONTH_COL, end_row=2, end_column=TS_BARRIERS_COL)
    except Exception:
        pass
    ws.row_dimensions[1].height = 25

    unmerge_row(ws, TS_HEADER_ROW)

    # Write section color fills across the TS header row
    for start, end, label, color in TS_SECTION_HEADERS:
        for c in range(start, end + 1):
            ws.cell(row=TS_HEADER_ROW, column=c).fill = PatternFill("solid", fgColor=color)

    # Cols 1-3 = blank spacer, TS_MONTH_COL = Months, TS_BARRIERS_COL = Total Barriers
    for c in range(1, OUT_DATA_START):
        cell = ws.cell(row=TS_HEADER_ROW, column=c)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
    for col, text in [
        (TS_MONTH_COL,    "Months"),
        (TS_BARRIERS_COL, "Total\nBarriers"),
    ]:
        cell = ws.cell(row=TS_HEADER_ROW, column=col, value=text)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type=None)

    # Data cols (6..MAX_OUT_COL) now write directly at their own out_col
    # position — cols 3-5 are left blank, taking the place of the 3
    # identity columns (Territory #/VCART Name/Total Accounts) that only
    # the LIVE/snapshot sections need. See TS_SECTION_HEADERS in config.py.
    for out_col in range(OUT_DATA_START, MAX_OUT_COL + 1):
        text = COL_HEADERS.get(out_col, "")
        cell = ws.cell(row=TS_HEADER_ROW, column=out_col, value=text)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = _section_fill(out_col, TS_SECTION_HEADERS)
        fill_color = next((c for s, e, _, c in TS_SECTION_HEADERS if s <= out_col <= e), None)
        text_color = "FFFFFF" if fill_color in VIIV_DARK_COLORS else "000000"
        cell.font = Font(bold=True, size=9, color=text_color)
    ws.row_dimensions[TS_HEADER_ROW].height = 80

    for r in range(TS_START_ROW, TS_END_ROW + 1):
        ws.row_dimensions[r].height = 16

    apply_column_widths(ws, force=True)


def apply_column_widths(ws, force=False):
    """Apply layout-specific column widths to the VCART Totals sheet.

    force=True (first_run / brand-new file): always set every width.
    force=False (every later run): only fill in a width for a column with
    NO width recorded at all — preserves manual resizing in Excel across
    runs while still backfilling any newly-added column that's missing a
    width entirely (e.g. cols 42-44 after this restructure). See ADFR's
    main.py for the fuller rationale (same pattern, ported here).
    """
    defaults = {"A": 25, "B": 18, "C": 16, "D": 16}
    for col in ["E", "F", "G", "H"]:
        defaults[col] = 10
    for col in ["I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        defaults[col] = 28
    for col in ["U", "V", "W", "X", "Y", "Z"]:
        defaults[col] = 11
    for idx in range(27, MAX_OUT_COL + 4):  # covers TS cols beyond old range too
        defaults[get_column_letter(idx)] = 11

    for letter, width in defaults.items():
        if force or ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = width


def write_live_section(ws, all_td, snap_date):
    o = get_live_offsets(ws)

    date_str = snap_date.strftime("%m.%d.%y").lstrip("0").replace(".0", ".")
    ws.cell(row=o["live_row"], column=1,
            value=f"LIVE - refreshed {date_str}").font = Font(size=11)
    ws.cell(row=o["live_row"], column=1).fill = PatternFill("solid", fgColor=YELLOW)

    total_barriers = compute_total_barriers(all_td)
    ws.cell(row=o["live_row"], column=2,
            value=f"Total barriers: {total_barriers}").font = Font(size=11)
    ws.cell(row=o["live_row"], column=2).fill = PatternFill("solid", fgColor=YELLOW)

    # RAW section
    write_section_headers(ws, o["raw_sec_row"])
    write_live_col_headers(ws, o["raw_hdr_row"])
    for i, (label, *_) in enumerate(TERRITORIES):
        write_territory_row(ws, o["raw_data_row"] + i, label, all_td.get(label), raw=True)
    write_nation_row(ws, o["raw_nat_row"], all_td, raw=True)
    ws.row_dimensions[o["gap_row"]].height = 8

    # PCT section
    ws.cell(row=o["pct_label_row"], column=1,
            value="% of LIVE").font = Font(size=11, bold=True)
    ws.cell(row=o["pct_label_row"], column=1).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[o["pct_label_row"]].height = 16
    write_section_headers(ws, o["pct_sec_row"])
    write_live_col_headers(ws, o["pct_hdr_row"])
    for i, (label, *_) in enumerate(TERRITORIES):
        write_territory_row(ws, o["pct_data_row"] + i, label, all_td.get(label), raw=False)
    write_nation_row(ws, o["pct_nat_row"], all_td, raw=False)
    ws.row_dimensions[o["pct_nat_row"] + 1].height = 8


# ---------------------------------------------------------------------------
# TIME SERIES
# ---------------------------------------------------------------------------

def _parse_ts_row_date(cell_val):
    """Extract (year, month) from a time-series row label like
    'January (01/15/26)'. Returns None if it can't be parsed.
    """
    m = re.search(r"\((\d{1,2})/(\d{1,2})/(\d{2,4})\)", str(cell_val or ""))
    if not m:
        return None
    mm, _dd, yy = m.groups()
    yy = int(yy)
    year = 2000 + yy if yy < 100 else yy
    return (year, int(mm))


def update_time_series(ws, all_td, snap_date):
    month_name = snap_date.strftime("%B")
    month_row  = None

    for r in range(TS_START_ROW, TS_END_ROW + 1):
        cell_val = ws.cell(row=r, column=TS_MONTH_COL).value
        if not cell_val or str(cell_val).strip() == "":
            month_row = r
            break
        # Match on (year, month), not just month NAME — see ADFR's
        # equivalent function for the full explanation of why a name-only
        # match silently corrupts row order once a month name recurs in a
        # later year.
        parsed = _parse_ts_row_date(cell_val)
        if parsed == (snap_date.year, snap_date.month):
            month_row = r
            break

    if not month_row:
        for r in range(TS_START_ROW, TS_END_ROW):
            for c in range(1, MAX_OUT_COL + 1):
                try:
                    src = ws.cell(row=r + 1, column=c)
                    dst = ws.cell(row=r, column=c)
                    dst.value = src.value
                    if src.has_style:
                        dst.font          = copy(src.font)
                        dst.fill          = copy(src.fill)
                        dst.alignment     = copy(src.alignment)
                        dst.number_format = src.number_format
                except Exception:
                    pass
        month_row = TS_END_ROW

    valid = [td for td in all_td.values() if td]
    total_barriers = compute_total_barriers(all_td)
    total_campuses = sum(td["campus_count"] for td in valid)

    unmerge_row(ws, month_row)
    # Clear cols 1..OUT_DATA_START-1 first — guards against stale content
    # left over in the old A/B position from before Months/Total Barriers
    # moved to D/E (see TS_MONTH_COL/TS_BARRIERS_COL), since this function
    # no longer touches cols 1-3 going forward.
    for c in range(1, OUT_DATA_START):
        cell = ws.cell(row=month_row, column=c)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
    date_label = f"{month_name} ({snap_date.strftime('%m/%d/%y')})"
    ws.cell(row=month_row, column=TS_MONTH_COL, value=date_label).font = Font(size=11)
    ws.cell(row=month_row, column=TS_MONTH_COL).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=month_row, column=TS_BARRIERS_COL, value=total_barriers).font = Font(size=11)
    ws.cell(row=month_row, column=TS_BARRIERS_COL).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[month_row].height = 16

    # Data now writes at each field's own out_col position directly —
    # cols 3-5 stay blank, same alignment change as write_totals_sheet_headers.
    for out_col in range(OUT_DATA_START, MAX_OUT_COL + 1):
        cell = ws.cell(row=month_row, column=out_col)
        cell.fill = PatternFill(fill_type=None)
        if out_col in BARRIER_COLS_OUT:
            nums = [td["totals"].get(out_col) for td in valid
                    if isinstance(td["totals"].get(out_col), (int, float))]
            if nums:
                cell.value = sum(nums)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        elif out_col in BARRIER_DAYS_TO_CLOSE_COLS:
            # Same pooled method as write_nation_row (see there for why):
            # pool every individual closed barrier nationwide (properly
            # quarter-windowed), Barrier 1 and Barrier 2 kept separate —
            # not a mode/most-common of each territory's own value (which
            # is what the generic else branch below would otherwise do,
            # since these cols aren't numeric sums). This row represents
            # the SAME current pull as the LIVE section's NATION row, so
            # the two must agree.
            key = ("barrier1_days_to_close_quarter" if out_col == BARRIER_DAYS_TO_CLOSE_COLS[0]
                   else "barrier2_days_to_close_quarter")
            pool = [d for td in valid for d in td["totals"].get(key, [])]
            if pool:
                cell.value = sum(pool) / len(pool)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"
        elif out_col in PCT_COLS_OUT:
            # Same weighted-average method as write_nation_row (see there
            # for why): sum(pct_i * campus_count_i) / total_campuses, not
            # a mode of each territory's own raw fraction — which is what
            # the generic else branch below would otherwise do, producing
            # e.g. "whichever fraction happened to repeat most often"
            # instead of the actual nationwide percentage. This row
            # represents the SAME current pull as the LIVE section's
            # NATION row, so the two must agree.
            weighted_sum = 0.0
            weight_total = 0
            for td in valid:
                v = td["totals"].get(out_col)
                if isinstance(v, (int, float)) and td["campus_count"]:
                    weighted_sum += v * td["campus_count"]
                    weight_total += td["campus_count"]
            if weight_total:
                cell.value = weighted_sum / weight_total
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0%"
        else:
            vals = [str(td["totals"].get(out_col) or "").strip() for td in valid
                    if td["totals"].get(out_col)]
            vals = [v for v in vals if v and v not in ("#N/A", "None")]
            if vals:
                cell.value = Counter(vals).most_common(1)[0][0]
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# ROW SHIFT HELPER
# ---------------------------------------------------------------------------

def _shift_rows_down(ws, start_row, end_row, delta):
    if delta <= 0 or end_row < start_row:
        return

    merges = [mr for mr in list(ws.merged_cells.ranges)
              if mr.min_row >= start_row and mr.max_row <= end_row]
    for mr in merges:
        ws.unmerge_cells(str(mr))

    for r in range(end_row, start_row - 1, -1):
        ws.row_dimensions[r + delta].height = ws.row_dimensions[r].height
        for c in range(1, MAX_OUT_COL + 1):
            src = ws.cell(row=r, column=c)
            dst = ws.cell(row=r + delta, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font          = copy(src.font)
                dst.fill          = copy(src.fill)
                dst.border        = copy(src.border)
                dst.alignment     = copy(src.alignment)
                dst.number_format = src.number_format
            src.value = None
            src.fill  = PatternFill(fill_type=None)
            src.font  = Font()
            src.border = Border()
            src.alignment = Alignment()
            src.number_format = "General"

    for mr in merges:
        ws.merge_cells(start_row=mr.min_row + delta, start_column=mr.min_col,
                       end_row=mr.max_row + delta, end_column=mr.max_col)


# ---------------------------------------------------------------------------
# SNAPSHOT MANAGEMENT
# ---------------------------------------------------------------------------

def freeze_live_as_snapshot(ws, live_month, snap_date):
    o = get_live_offsets(ws)
    live_start = o["live_row"]
    live_end   = o["pct_nat_row"] + 1
    block_size = live_end - live_start + 1
    pct_label_offset = o["pct_label_row"] - live_start

    snap_start = get_snap_start(ws)

    _shift_rows_down(ws, snap_start, ws.max_row, block_size)

    live_merges = [mr for mr in list(ws.merged_cells.ranges)
                   if live_start <= mr.min_row <= live_end]
    for mr in live_merges:
        ws.unmerge_cells(str(mr))

    offset = snap_start - live_start
    for i in range(block_size):
        src_row = live_start + i
        dst_row = snap_start + i
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for c in range(1, MAX_OUT_COL + 1):
            src = ws.cell(row=src_row, column=c)
            dst = ws.cell(row=dst_row, column=c)
            dst.value = src.value
            if src.has_style:
                dst.font          = copy(src.font)
                dst.fill          = copy(src.fill)
                dst.border        = copy(src.border)
                dst.alignment     = copy(src.alignment)
                dst.number_format = src.number_format
            src.value = None
            src.fill  = PatternFill(fill_type=None)
            src.font  = Font()
            src.border = Border()
            src.alignment = Alignment()
            src.number_format = "General"

    for mr in live_merges:
        ws.merge_cells(start_row=mr.min_row + offset, start_column=mr.min_col,
                       end_row=mr.max_row + offset, end_column=mr.max_col)

    old_label_text = str(ws.cell(row=snap_start, column=1).value or "")
    date_match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", old_label_text)
    if date_match:
        mm, dd, yy = date_match.groups()
        date_str = f"{int(mm):02d}/{int(dd):02d}/{yy[-2:].zfill(2)}"
    else:
        date_str = snap_date.strftime("%m/%d/%y")
    ws.cell(row=snap_start, column=1, value=f"{live_month} snapshot ({date_str})")
    barriers_val = ws.cell(row=snap_start, column=2).value
    m = re.search(r"(\d+)", str(barriers_val) or "")
    if m:
        ws.cell(row=snap_start, column=2, value=f"Total barriers - {m.group(1)}")

    pct_label_cell = ws.cell(row=snap_start + pct_label_offset, column=1)
    pct_label_cell.value = None
    pct_label_cell.fill = PatternFill(fill_type=None)

    _trim_snapshots(ws, keep=11)


def _trim_snapshots(ws, keep=11):
    """Drop oldest snapshot blocks beyond `keep`, shifting the remaining
    (newer) ones up to fill the gap.

    FIX: this used to move only cell .value when shifting rows up, leaving
    font/fill/border/alignment/number_format behind at the old row index.
    In particular the snapshot label row's yellow fill (what is_yellow_row /
    find_yellow_rows use to locate snapshot boundaries on every subsequent
    run) stayed put instead of following its label text up — so after a
    trim, the shifted label row would read correctly but no longer be
    detected as "yellow", corrupting boundary detection on the next trim.
    Now copies full cell style, same pattern as _shift_rows_down /
    freeze_live_as_snapshot, so the whole cell (value + formatting) moves
    together.
    """
    snap_start = get_snap_start(ws)
    snap_rows  = find_yellow_rows(ws, snap_start, ws.max_row)
    if len(snap_rows) <= keep:
        return
    to_delete    = len(snap_rows) - keep
    delete_up_to = snap_rows[to_delete] - 2
    for r in range(snap_start, delete_up_to + 1):
        for c in range(1, MAX_OUT_COL + 1):
            try:
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font()
                cell.border = Border()
                cell.alignment = Alignment()
                cell.number_format = "General"
            except Exception:
                pass
    remaining, shift_to = delete_up_to + 1, snap_start
    while remaining <= ws.max_row:
        ws.row_dimensions[shift_to].height = ws.row_dimensions[remaining].height
        for c in range(1, MAX_OUT_COL + 1):
            try:
                src = ws.cell(row=remaining, column=c)
                dst = ws.cell(row=shift_to, column=c)
                dst.value = src.value
                if src.has_style:
                    dst.font          = copy(src.font)
                    dst.fill          = copy(src.fill)
                    dst.border        = copy(src.border)
                    dst.alignment     = copy(src.alignment)
                    dst.number_format = src.number_format
                src.value = None
                src.fill  = PatternFill(fill_type=None)
                src.font  = Font()
                src.border = Border()
                src.alignment = Alignment()
                src.number_format = "General"
            except Exception:
                pass
        shift_to += 1
        remaining += 1


# ---------------------------------------------------------------------------
# LIVE MONTH DETECTION
# ---------------------------------------------------------------------------

def detect_live_month(ws):
    """Return (month_name, year, month_num) for the LIVE label's date, or
    None.

    FIX: this used to scan every row in column 1 for the substring "live"
    (`if "live" in v.lower()`), rather than reading the LIVE label at its
    known, deterministic position. That happened to work only because
    nothing else in column 1 currently contains "live" — a future
    territory/label/region name containing that substring (e.g. anything
    with "olive", "livelihood", etc., however unlikely today) would have
    matched the wrong row silently. The LIVE row's position is already
    computed deterministically by get_live_offsets(), so read that cell
    directly instead of scanning the whole column.

    Year-aware for the same reason as ADFR's version — a name-only
    comparison would misread a gap of exactly 12+ months landing on the
    same calendar month as "same month, refresh in place", silently
    skipping the snapshot freeze.
    """
    live_row = get_live_offsets(ws)["live_row"]
    if live_row > ws.max_row:
        return None
    v = str(ws.cell(row=live_row, column=1).value or "")
    if "live" not in v.lower():
        return None
    date_match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", v)
    if date_match:
        mm, _dd, yy = date_match.groups()
        month_num = int(mm)
        if 1 <= month_num <= 12:
            yy = int(yy)
            year = 2000 + yy if yy < 100 else yy
            return (MONTH_NAMES[month_num - 1], year, month_num)
    return None


# ---------------------------------------------------------------------------
# VCART SYSTEMS — LIVE sheet (campus-level)
# ---------------------------------------------------------------------------

def _hide_barrier_date_columns(ws):
    """Hide Barrier 1/2 Start Date and Close Date columns (N, O, S, T) on
    the VCART Totals sheet per request. Days to Close (P, U) stays visible
    — only the raw date columns hide.
    """
    for out_col in HIDDEN_BARRIER_DATE_COLS:
        ws.column_dimensions[get_column_letter(out_col)].hidden = True


def _hide_systems_barrier_date_columns(ws):
    """Hide Barrier 1/2 Start Date and Close Date columns (Q, R, V, W) on
    'VCART Systems - LIVE' per request. Days to Close (S, X) stays visible
    — only the raw date columns hide. Column-hidden state travels with the
    worksheet when it's later renamed into a historical snapshot, so this
    only needs to run when a fresh LIVE sheet is built each month.
    """
    for col in HIDDEN_SYSTEMS_BARRIER_DATE_COLS:
        ws.column_dimensions[get_column_letter(col)].hidden = True


def _position_trends_sheet(wb):
    """Place 'Trends & Takeaways' immediately after 'VCART Systems - LIVE'
    (before the historical 'VCART Systems - <Mon Year>' snapshot sheets),
    per request. build_summary_sheet() always deletes+recreates this sheet
    at the end of the workbook, so this re-fixes its position every run
    regardless of how many historical snapshot sheets have accumulated.
    """
    if "Trends & Takeaways" not in wb.sheetnames or "VCART Systems - LIVE" not in wb.sheetnames:
        return
    target_index = wb.sheetnames.index("VCART Systems - LIVE") + 1
    current_index = wb.sheetnames.index("Trends & Takeaways")
    if current_index != target_index:
        wb.move_sheet("Trends & Takeaways", offset=target_index - current_index)


def _position_systems_live_sheet(wb):
    if "VCART Systems - LIVE" not in wb.sheetnames:
        return
    target_index = 1
    current_index = wb.sheetnames.index("VCART Systems - LIVE")
    if current_index != target_index:
        wb.move_sheet("VCART Systems - LIVE", offset=target_index - current_index)


def build_systems_live_sheet(wb, all_td, snap_date):
    """Write campus-level data to 'VCART Systems - LIVE' sheet."""
    if "VCART Systems - LIVE" in wb.sheetnames:
        del wb["VCART Systems - LIVE"]
    ws = wb.create_sheet("VCART Systems - LIVE")

    title_cell = ws.cell(row=1, column=1, value=f"VCART Systems - LIVE  (refreshed {snap_date.strftime('%m/%d/%y')})")
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_campuses = sum(td["campus_count"] for td in all_td.values() if td)
    nation_barriers = [
        sum(td["totals"].get(c, 0) or 0 for td in all_td.values()
            if td and isinstance(td["totals"].get(c), (int, float)))
        for c in BARRIER_COLS_OUT
    ]
    totals_label_row = 2
    ws.cell(row=totals_label_row, column=8, value="TOTALS").font = Font(bold=True, size=9)
    ws.cell(row=totals_label_row, column=8).fill = PatternFill("solid", fgColor=NATION_FILL)
    for i, bc in enumerate(BARRIER_COLS_OUT):
        out_col_offset = bc - OUT_DATA_START
        ws.cell(row=totals_label_row, column=9 + out_col_offset,
                value=nation_barriers[i]).font = Font(size=9, bold=True)

    total_barriers_col = 9 + len(BARRIER_COLS_OUT)
    total_barriers_cell = ws.cell(row=totals_label_row, column=total_barriers_col,
                                   value=f"Total Barriers: {sum(nation_barriers)}")
    total_barriers_cell.font = Font(bold=True, size=9)
    total_barriers_cell.fill = PatternFill("solid", fgColor=NATION_FILL)

    write_section_headers(ws, 3, headers=SYSTEMS_SECTION_HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=7)

    campus_hdrs = {
        1: "VCART Territory #",
        2: "VCART Name",
        3: "Corporate\nParent CID",
        4: "Corporate\nParent Name",
        5: "Address",
        6: "City",
        7: "State",
        8: "Zip",
    }
    for col, text in campus_hdrs.items():
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = Font(size=9, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for out_col, text in COL_HEADERS.items():
        if out_col >= OUT_DATA_START:
            src_col = SRC_DATA_START + (out_col - OUT_DATA_START)
            ws_col  = src_col
            cell = ws.cell(row=4, column=ws_col, value=text)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = _section_fill(out_col)
            fill_color = next((c for s, e, _, c in SECTION_HEADERS if s <= out_col <= e), None)
            text_color = "FFFFFF" if fill_color in VIIV_DARK_COLORS else "000000"
            cell.font = Font(size=9, color=text_color)
    ws.row_dimensions[4].height = 80

    current_row = 5
    for label, filename in TERRITORIES:
        td = all_td.get(label)
        if td is None or not td["campus_rows"]:
            continue
        for campus in td["campus_rows"]:
            def _demog(col, val):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = Font(size=10)
                is_number = isinstance(val, (int, float))
                cell.alignment = Alignment(horizontal="center" if is_number else "left",
                                            vertical="center")
            _demog(1, campus.get("terr_num", td["terr_num"]))
            _demog(2, campus.get("vcart_name", td["vcart_name"]))
            _demog(3, campus.get("cid"))
            _demog(4, campus.get("corp_name"))
            _demog(5, campus.get("address"))
            _demog(6, campus.get("city"))
            _demog(7, campus.get("state"))
            _demog(8, campus.get("zip"))

            for out_col, val in campus.items():
                if not isinstance(out_col, int):
                    continue
                src_offset = out_col - OUT_DATA_START
                ws_col = SRC_DATA_START + src_offset
                cell = ws.cell(row=current_row, column=ws_col, value=val)
                cell.font = Font(size=10)
                is_number = isinstance(val, (int, float))
                cell.alignment = Alignment(horizontal="center" if is_number else "left",
                                            vertical="center")
            ws.row_dimensions[current_row].height = 14
            current_row += 1

    _autofit_column_widths(ws)

    print(f"    VCART Systems - LIVE: {current_row - 5} campus rows written")


def rename_live_sheet_as_snapshot(wb, live_month, snap_date):
    if "VCART Systems - LIVE" not in wb.sheetnames:
        return

    ws_live = wb["VCART Systems - LIVE"]
    sheet_name = f"VCART Systems - {live_month[:3]} {snap_date.strftime('%Y')}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    old_title = ws_live.cell(row=1, column=1).value or ""
    new_title = re.sub(
        r"VCART Systems - LIVE\s*\(refreshed [^)]*\)",
        f"VCART Systems Snapshot — {live_month} {snap_date.strftime('%Y')}",
        old_title,
    )
    ws_live.cell(row=1, column=1).value = new_title
    ws_live.title = sheet_name


# ---------------------------------------------------------------------------
# FINAL FORMATTING PASS
# ---------------------------------------------------------------------------

def finalize_workbook_formatting(wb):
    AB_COLS = (1, 2)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

        # "Trends & Takeaways" already sets its own borders per table (see
        # the bd() helpers in summary.py), each scoped to that table's real
        # column width (3, 6, 8, 9 cols, etc.) within a wider max_col=20
        # canvas. This pass borders every column out to max_col on any row
        # that has data anywhere — on this sheet that means blank trailing
        # cells past a table's real width get boxed too, reading as an
        # unwanted grid. Skip it here; only sheets that genuinely use their
        # full column width per row (VCART Totals, the Systems sheets) rely
        # on this pass for borders.
        if ws.title == "Trends & Takeaways":
            continue

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue

        # On "VCART Totals", cols A-C above the LIVE row are the blank
        # spacer for the top TS table (Months/Total Barriers sit at D/E) —
        # never border those specific cells, even on a row where D onward
        # has real data, so the spacer doesn't get boxed in for no reason.
        # Cleared explicitly (not just skipped) so a border already baked
        # in from a prior run — before this exemption existed — is removed
        # on the next run rather than left sitting there.
        ts_spacer_end_row = None
        if ws.title == "VCART Totals":
            ts_spacer_end_row = get_live_offsets(ws)["live_row"]

        if ts_spacer_end_row is not None:
            # "VCART Totals": the title lives at D1 (TS_MONTH_COL) now, so
            # col 1 is always blank — the old a1_val check below silently
            # stopped doing anything once the title moved off A1, and
            # never cleared a border already sitting on B1/C1 from before
            # that move. Clear cols 1-3 explicitly, same treatment as the
            # rest of the spacer zone above LIVE. Never boxes the title
            # itself (D1 onward) — that was never bordered before the
            # move, and nothing here should start boxing it now.
            for c in (1, 2, 3):
                cell = ws.cell(row=1, column=c)
                if cell.border and cell.border.left and cell.border.left.style:
                    cell.border = Border()
        else:
            a1_val = ws.cell(row=1, column=1).value
            if a1_val is not None and str(a1_val).strip() != "":
                ws.cell(row=1, column=1).border = ALL_BORDERS

        for r in range(2, max_row + 1):
            has_data = any(
                ws.cell(row=r, column=c).value is not None and
                str(ws.cell(row=r, column=c).value).strip() != ""
                for c in range(1, max_col + 1)
            )
            if not has_data:
                continue

            in_ts_spacer_zone = ts_spacer_end_row is not None and r < ts_spacer_end_row

            if is_yellow_row(ws, r):
                for c in AB_COLS:
                    if c <= max_col:
                        if in_ts_spacer_zone and c <= 3:
                            ws.cell(row=r, column=c).border = Border()
                        else:
                            ws.cell(row=r, column=c).border = ALL_BORDERS
                continue

            for c in range(1, max_col + 1):
                if in_ts_spacer_zone and c <= 3:
                    ws.cell(row=r, column=c).border = Border()
                    continue
                ws.cell(row=r, column=c).border = ALL_BORDERS


def main():
    t0 = datetime.now()
    print("=" * 60)
    print(f"  VCART VRHR Aggregator  —  {t0.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    print(f"\nReading territory files from:\n  {BASE_DIR}\n")
    all_td = load_all_territories(BASE_DIR)

    print("\nLoading barrier taxonomy (category <-> subcategory)...")
    taxonomy = load_barrier_taxonomy(BASE_DIR)

    loaded = sum(1 for td in all_td.values() if td)
    total_campuses = sum(td["campus_count"] for td in all_td.values() if td)
    print(f"\n  {loaded}/{NUM_TERRITORIES} territories loaded  |  {total_campuses} total campuses")

    if loaded < NUM_TERRITORIES:
        failed = [l for l, td in all_td.items() if not td]
        print(f"\n  WARNING: Failed to load: {', '.join(failed)}")
        answer = input("\n  Continue anyway? (y/n): ").strip().lower()
        if answer != "y":
            print("\n  Aborted.")
            return

    snap_date     = datetime.now()
    current_month = snap_date.strftime("%B")

    print(f"\nChecking for existing output file...")
    first_run = not os.path.exists(OUTPUT)
    if first_run:
        print(f"  Not found — this is a first run.")
        wb = Workbook()
        ws_totals = wb.active
        ws_totals.title = "VCART Totals"
        write_totals_sheet_headers(ws_totals)
        live_info = None
    else:
        print(f"  Found: {os.path.basename(OUTPUT)}")
        wb = load_workbook(OUTPUT)
        ws_totals = wb["VCART Totals"]
        live_info = detect_live_month(ws_totals)

    live_month = live_info[0] if live_info else None
    save_snapshot = bool(live_info and (live_info[1], live_info[2]) != (snap_date.year, snap_date.month))
    if save_snapshot:
        print(f"\n  Month rollover detected: {live_month} → {current_month}")
        print(f"  Will freeze {live_month}'s LIVE section into a snapshot before writing {current_month} data.")
    elif live_month:
        print(f"\n  Same month ({current_month}). Refreshing LIVE data only.")
    else:
        print(f"\n  First run. Writing fresh file.")

    print("\n" + "-" * 60)
    print("  Updating workbook in place...")
    print("-" * 60)

    old_live_row = None if first_run else get_live_offsets(ws_totals)["live_row"]

    print(f"\n  [1/3] Writing {current_month} to time series...")
    update_time_series(ws_totals, all_td, snap_date)

    if not first_run:
        new_live_row = get_live_offsets(ws_totals)["live_row"]
        delta = new_live_row - old_live_row
        if delta > 0:
            print(f"        Time series grew — shifting LIVE section and snapshots down {delta} row(s).")
            _shift_rows_down(ws_totals, old_live_row, ws_totals.max_row, delta)

    total_barriers = compute_total_barriers(all_td)
    print(f"        {current_month} written — {total_barriers} total barriers, {total_campuses} campuses.")

    if save_snapshot:
        print(f"\n  [2/3] Freezing {live_month}'s LIVE section into a snapshot...")
        freeze_live_as_snapshot(ws_totals, live_month, snap_date)
        print(f"        {live_month} snapshot written.")
    else:
        print(f"\n  [2/3] No rollover — LIVE section will be refreshed in place.")

    print(f"\n  [2/3] Writing LIVE section ({NUM_TERRITORIES} territories + NATION, raw + %)...")
    write_live_section(ws_totals, all_td, snap_date)
    print(f"        LIVE section done.")

    if save_snapshot and not first_run:
        print(f"\n  [3/3] Renaming VCART Systems - LIVE → snapshot for {live_month}...")
        rename_live_sheet_as_snapshot(wb, live_month, snap_date)

    print(f"\n  [3/3] Building VCART Systems - LIVE sheet (campus-level rows)...")
    build_systems_live_sheet(wb, all_td, snap_date)
    _position_systems_live_sheet(wb)
    _hide_systems_barrier_date_columns(wb["VCART Systems - LIVE"])

    print(f"\n  [4/4] Building Trends & Takeaways...")
    summary.build_summary_sheet(wb, all_td, ws_totals, taxonomy)
    _position_trends_sheet(wb)

    _hide_barrier_date_columns(ws_totals)

    print(f"\n" + "-" * 60)
    print(f"  Saving → {OUTPUT}")
    print("-" * 60)
    apply_column_widths(ws_totals)
    finalize_workbook_formatting(wb)
    wb.save(OUTPUT)

    elapsed = (datetime.now() - t0).seconds
    print(f"\n{'=' * 60}")
    print(f"  DONE in {elapsed}s")
    print(f"  {loaded}/{NUM_TERRITORIES} territories  |  {total_campuses} campuses  |  {total_barriers} barriers")
    print(f"  Output: {os.path.basename(OUTPUT)}")
    print(f"{'=' * 60}")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        print(traceback.format_exc())
        input("\nPress Enter to exit...")