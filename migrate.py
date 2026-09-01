"""
migrate.py — VCART
====================
ONE-TIME migration for the VCART Totals output workbook's time-series
table (row 3 header + month rows), moving existing historical data from
the OLD column layout to the NEW layout introduced by this round of
fixes.

Frozen snapshot blocks are deliberately NOT migrated — see "What this
does NOT touch" below.

WHAT CHANGED, AND WHY EVERY TS COLUMN SHIFTS (NOT JUST THE ONES AFTER
BARRIER TRACKING)
-------------------------------------------------------------------------
1. The TS table used to place its data 3 columns to the LEFT of where
   the exact same field sits in the LIVE section below it (the TS table
   only needs 2 label columns — Month, Total Barriers — instead of the
   5 identity columns LIVE needs, so its data started at column 3
   instead of column 6). That meant the same column letter meant a
   DIFFERENT field depending which table you were looking at — column N
   was "Barrier 1 Start Date" in LIVE but something else entirely in the
   TS table above it. This migration removes that offset entirely: TS
   data now starts at the same column as everywhere else, with cols 3-5
   left as blank spacers. This shifts EVERY TS data column, including
   the Implementation Barrier percentages, which hadn't otherwise
   changed at all.

2. Two new "Days to Close" columns were added (one after each Barrier's
   Close Date). No historical equivalent ever existed, so these are left
   blank going back — nothing to migrate, they'll start populating from
   the next real run onward.

3. Barrier 2's fields (Category/Subcategory/Start Date/Close Date) and
   both Implementation Stage fields shift further to make room for the
   new Days-to-Close columns. This part IS safely migrated — that data
   was already correctly header-matched in the old script, just sitting
   at different column numbers.

4. Everything from the old "CARE Team CEP Stage" column onward was read
   POSITIONALLY in the old script, and this was a real, active bug — two
   real territory files in the same quarter don't even use the same
   column layout for this section, so the old script was silently
   reading different, unrelated fields into the same output column
   depending on which territory's row it happened to be reading.
   Confirmed directly against a real output file: the column labeled
   "CARE Team CEP Stage" held the exact same value as the adjacent
   "Current Impl. Stage" column in every row checked — that field never
   existed; the old script was reading stray data into it.

   Because different territories were misaligned in DIFFERENT ways, the
   NATION-level sums/percentages/modes in this zone blend together
   readings that don't correspond to the same real field across
   territories. There's no reliable way to un-scramble that after the
   fact. Rather than migrate wrong data under new, correctly-labeled
   columns, this migration leaves the new Territory Fields columns
   blank for all historical data. Every run from here on populates them
   correctly.

WHAT THIS DOES NOT TOUCH
-------------------------------------------------------------------------
Frozen snapshot blocks (the dated historical detail blocks below the
LIVE section) are left completely as they were — not migrated, not
reformatted, not touched at all. Migrating them accurately would mean
tracking two different column-alignment schemes at once for something
that's only ever used for point-in-time reference, not active trend
analysis — not worth the complexity. They'll simply keep aging out
naturally as new snapshots get frozen going forward (max 11 kept,
oldest trimmed automatically), so this isn't a permanent state — it's
just how the transition period looks.

Does NOT touch the source file. Writes a new file
(<original_name>.MIGRATED.xlsx) so you can review before replacing the
original.

Usage:
    python migrate.py "C:\\path\\to\\VCART.Region.Nation.Totals.New.xlsx"
"""

import sys
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import (
    COL_HEADERS, SECTION_HEADERS, OUT_DATA_START,
    TS_HEADER_ROW, TS_START_ROW, TS_END_ROW, MONTH_NAMES,
    VIIV_DARK_COLORS,
)

# old_out_col -> new_out_col, for the part of the layout that's safely
# migratable (identity + Implementation Barriers unchanged; Barrier
# Tracking's Category/Subcategory/Start/Close/Stage fields shifted to
# make room for the two new Days-to-Close columns). Old cols 22-44 have
# NO entry here — deliberately dropped, see module docstring.
OLD_TO_NEW_COL = {}
for c in range(1, 12):        # identity (1-5) + Implementation Barriers (6-11): unchanged
    OLD_TO_NEW_COL[c] = c
OLD_TO_NEW_COL[12] = 12   # Barrier 1 Category
OLD_TO_NEW_COL[13] = 13   # Barrier 1 Subcategory
OLD_TO_NEW_COL[14] = 14   # Barrier 1 Start Date
OLD_TO_NEW_COL[15] = 15   # Barrier 1 Close Date
# (new col 16 = Barrier 1 Days to Close — no old equivalent)
OLD_TO_NEW_COL[16] = 17   # Barrier 2 Category
OLD_TO_NEW_COL[17] = 18   # Barrier 2 Subcategory
OLD_TO_NEW_COL[18] = 19   # Barrier 2 Start Date
OLD_TO_NEW_COL[19] = 20   # Barrier 2 Close Date
# (new col 21 = Barrier 2 Days to Close — no old equivalent)
OLD_TO_NEW_COL[20] = 22   # Start of Qtr Impl. Stage
OLD_TO_NEW_COL[21] = 23   # Current Impl. Stage
# old cols 22-44 intentionally have no mapping — dropped, see docstring.

OLD_MAX_OUT_COL = 44


def _old_out_to_ts_col(old_out_col):
    """The OLD file's own TS-column formula — this file was written by
    the old script, which always used this -3 offset. Only ever applied
    to the OLD (source) side of the migration; the NEW side no longer
    uses any offset at all (see module docstring, point 1)."""
    return old_out_col - OUT_DATA_START + 3


# old_ts_col -> new_out_col directly (no more offset on the new side).
# Every out_col >= OUT_DATA_START shifts, including the Implementation
# Barrier percentages (6-11), which hadn't otherwise changed — old
# ts_col 3-8 moves to new out_col 6-11, a straight +3, purely from
# removing the offset.
OLD_TO_NEW_TS_COL = {
    _old_out_to_ts_col(old_out_col): new_out_col
    for old_out_col, new_out_col in OLD_TO_NEW_COL.items()
    if old_out_col >= OUT_DATA_START
}
OLD_MAX_TS_COL = _old_out_to_ts_col(OLD_MAX_OUT_COL)


def _migrate_row_cols(ws, row, col_map, max_old_col, min_col=1):
    """
    Remap one row's columns per col_map (old_col -> new_col), reading
    every old value into memory FIRST before writing anything — several
    columns shift by different amounts, so writing in place risks
    clobbering a value before it's been read. Columns from min_col to
    max_old_col that aren't in col_map (the dropped old CEP-onward zone)
    are cleared. Columns before min_col are left completely untouched —
    used for the TS table's cols 1-2 ("Months" / "Total Barriers"), which
    aren't part of col_map's remapping at all and must survive as-is.
    """
    old_vals = {}
    for c in range(min_col, max_old_col + 1):
        cell = ws.cell(row=row, column=c)
        old_vals[c] = (cell.value, copy(cell.font) if cell.has_style else None,
                        copy(cell.fill) if cell.has_style else None,
                        copy(cell.alignment) if cell.has_style else None,
                        cell.number_format)

    # Clear every column in the row first (both mapped and dropped source
    # positions may overlap with destination positions).
    for c in range(min_col, max_old_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font()
        cell.alignment = Alignment()
        cell.number_format = "General"

    for old_c, new_c in col_map.items():
        if old_c < min_col:
            continue
        val, font, fill, align, numfmt = old_vals[old_c]
        dst = ws.cell(row=row, column=new_c)
        dst.value = val
        if font is not None:
            dst.font = font
            dst.fill = fill
            dst.alignment = align
            dst.number_format = numfmt


def migrate_ts_header_row(ws):
    """Rewrite row 3 (the time-series header) to the new column labels at
    their new (unshifted) positions, leaving cols 1-2 ("Months" / "Total
    Barriers") untouched. Cols 3-5 are cleared and left as blank spacers.
    """
    # Clear the entire old TS data range first — old col 3 held real
    # barrier-percentage data (the old ts_col=3 offset), which is not
    # where col 3 belongs in the new layout (a blank spacer). Clearing
    # everything from col 3 through the old file's own max column avoids
    # leaving stale old-position data sitting underneath the new labels.
    for col in range(3, OLD_MAX_TS_COL + 5):
        c = ws.cell(row=TS_HEADER_ROW, column=col)
        if c.value is not None:
            c.value = None
        c.fill = PatternFill(fill_type=None)

    for out_col, text in COL_HEADERS.items():
        if out_col < OUT_DATA_START:
            continue
        cell = ws.cell(row=TS_HEADER_ROW, column=out_col, value=text)
        fill_color = next((c for s, e, _, c in SECTION_HEADERS if s <= out_col <= e), None)
        text_color = "FFFFFF" if fill_color in VIIV_DARK_COLORS else "000000"
        cell.font = Font(bold=True, size=9, color=text_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=fill_color) if fill_color else PatternFill(fill_type=None)


def migrate_workbook(path):
    wb = load_workbook(path)
    ws = wb["VCART Totals"]

    print("  Migrating time-series header (row 3)...")
    migrate_ts_header_row(ws)

    print("  Migrating time-series month rows...")
    migrated_months = 0
    for r in range(TS_START_ROW, TS_END_ROW + 1):
        val = ws.cell(row=r, column=1).value
        if val and any(m.lower() in str(val).lower() for m in MONTH_NAMES):
            _migrate_row_cols(ws, r, OLD_TO_NEW_TS_COL, OLD_MAX_TS_COL, min_col=3)
            migrated_months += 1
    print(f"    {migrated_months} month row(s) migrated.")

    print("  Frozen snapshot blocks are left untouched — not migrated. See")
    print("  this script's module docstring for why.")

    out_path = path.rsplit(".", 1)[0] + ".MIGRATED.xlsx"
    wb.save(out_path)
    print(f"\n  Saved migrated copy -> {out_path}")
    print("\n  NOTE: columns 24-44 (Centralized/Decentralized through Last Update) are")
    print("  blank in all historical rows above. That data was never reliable in the")
    print("  old file — see this script's module docstring for why. Every run from")
    print("  here on populates those columns correctly. (The old 'Evolved' column has")
    print("  also been dropped entirely — it's a hidden field that was never meant")
    print("  to be used.)")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python migrate.py <path-to-existing-output.xlsx>")
        sys.exit(1)
    migrate_workbook(sys.argv[1])