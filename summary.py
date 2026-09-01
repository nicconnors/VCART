"""
Builds the "Trends & Takeaways" sheet: KEY METRICS cards, a monthly
Barrier Trends chart, category/subcategory breakdown tables (using the
Lists-sheet taxonomy), closed-barrier tracking (month + quarter), and
Days to Close by category (month + quarter).

"""

from collections import Counter
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from config import (
    TERRITORIES,
    BARRIER_COLS_OUT,
    VIIV_NAVY_DARKEST, VIIV_NAVY_MED, VIIV_MAROON_MED, VIIV_MAROON_DARK,
    MONTH_NAMES, TS_START_ROW, TS_END_ROW,
    OUT_DATA_START,
)

BARRIER_NAMES = ["Access for All / Coverage Challenges", "Financial & Reimbursement", "Knowledge & Training",
                  "Operational & Infra.", "Stakeholder Misalign.", "Technology & Data"]

# Hand-wrapped for the narrow-column trend/breakdown table headers below —
# blindly replacing every space with a newline (n.replace(" ", "\n")) is
# fine for short names, but on the two longer names above it produces an
# unreadably tall header, and at these columns' width Excel additionally
# wraps mid-word ("Reimbursement" splitting into "Reimbursemen"/"t") on
# top of the forced breaks. Wrapped by hand at natural word boundaries
# instead, same style as the COL_HEADERS entries in config.py. Keep in
# sync with BARRIER_NAMES above (same order, same 6 categories) — only
# the line breaks differ.
BARRIER_NAMES_WRAPPED = [
    "Access for All /\nCoverage\nChallenges",
    "Financial &\nReimbursement",
    "Knowledge &\nTraining",
    "Operational &\nInfra.",
    "Stakeholder\nMisalign.",
    "Technology &\nData",
]

BARRIER_CAT_OUT_COLS = {"Barrier 1": 12, "Barrier 2": 17}
BARRIER_SUBCAT_OUT_COLS = {"Barrier 1": 13, "Barrier 2": 18}

# Mirrors TS_MONTH_COL/TS_BARRIERS_COL in main.py — the top TS table's
# "Months"/"Total Barriers" labels sit at D/E now (cols 1-3 are a blank
# spacer). Derived the same way here so this reads from the same place
# main.py writes to, instead of the old A/B position.
TS_MONTH_COL    = OUT_DATA_START - 2   # 4 (D)
TS_BARRIERS_COL = OUT_DATA_START - 1   # 5 (E)


def _normalize_header(text):
    import re
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def tally_categories(campus_rows, cat_out_col):
    """Count how many campus rows have each category value in the given
    output column (Barrier 1 Category=12 or Barrier 2 Category=17)."""
    counts = Counter()
    for c in campus_rows:
        v = c.get(cat_out_col)
        if v and isinstance(v, str) and v.strip():
            counts[v.strip()] += 1
    return counts


def aggregate_category_counts(all_td, cat_out_col):
    """Sum category counts across every territory's campus_rows."""
    totals = Counter()
    for td in all_td.values():
        if not td:
            continue
        totals.update(tally_categories(td["campus_rows"], cat_out_col))
    return totals


def build_days_to_close_table(ws, r, all_td, live_tag=""):
    """
    Writes DAYS TO CLOSE BY CATEGORY starting at row r, stacked: once
    scoped to barriers closed THIS MONTH, then below it with no date
    filter (closed_count_quarter / category_days_to_close_quarter).
    Sorted worst (slowest) first. Returns the row after the second table.
    """
    DARK_BLUE, MED_BLUE = VIIV_NAVY_DARKEST, VIIV_NAVY_MED
    LIGHT_GREY, WHITE = "F2F2F2", "FFFFFF"

    def sh(cell, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    def bd(cell):
        s = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    def _write_table(start_row, title, data_key):
        cat_days = {}
        for td in all_td.values():
            if not td:
                continue
            for cat, days in td["totals"].get(data_key, []):
                cat_days.setdefault(cat, []).append(days)

        rr = start_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        title_cell = ws.cell(row=rr, column=1)
        sh(title_cell, bg=DARK_BLUE, size=12)
        title_cell.value = title
        ws.row_dimensions[rr].height = 24
        rr += 1

        for ci, h in enumerate(["Category", "Closed Count", "Avg Days to Close"], 1):
            c = ws.cell(row=rr, column=ci, value=h)
            sh(c, bg=MED_BLUE, size=9)
            bd(c)
        rr += 1

        ranked = sorted(cat_days.items(), key=lambda kv: -(sum(kv[1]) / len(kv[1])))
        for cat, days_list in ranked:
            avg_days = sum(days_list) / len(days_list)
            row_vals = [cat, len(days_list), avg_days]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=rr, column=ci, value=val)
                cell.font = Font(size=10, bold=(ci == 1))
                cell.alignment = Alignment(vertical="center", horizontal="left" if ci == 1 else "center")
                bd(cell)
                if ci == 3:
                    cell.number_format = "0"
                if rr % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            rr += 1
        return rr

    end_row_month   = _write_table(r, "DAYS TO CLOSE BY CATEGORY (Barrier 1 + Barrier 2 combined, closed this month only)" + live_tag,
                                    "category_days_to_close")
    end_row_quarter = _write_table(end_row_month + 1, "DAYS TO CLOSE BY CATEGORY (Barrier 1 + Barrier 2 combined, closed this quarter)" + live_tag,
                                    "category_days_to_close_quarter")
    return end_row_quarter + 1


def build_closed_barriers_table(ws, r, all_td, live_tag=""):
    """
    Writes CLOSED BARRIERS BY TERRITORY, stacked: Month above Quarter.
    Returns the row after the second table.
    """
    DARK_BLUE, MED_BLUE = VIIV_NAVY_DARKEST, VIIV_NAVY_MED
    LIGHT_GREY, WHITE = "F2F2F2", "FFFFFF"

    def sh(cell, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    def bd(cell):
        s = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    def _write_table(start_row, title, data_key):
        rr = start_row
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        sh(ws.cell(row=rr, column=1), bg=DARK_BLUE, size=12)
        ws.cell(row=rr, column=1).value = title
        ws.row_dimensions[rr].height = 24
        rr += 1

        for ci, h in enumerate(["Territory", "Closed Barriers"], 1):
            c = ws.cell(row=rr, column=ci, value=h)
            sh(c, bg=MED_BLUE, size=9)
            bd(c)
        rr += 1

        total = 0
        for label, *_ in TERRITORIES:
            td = all_td.get(label)
            closed = td["totals"].get(data_key, 0) if td else 0
            total += closed
            for ci, val in enumerate([label, closed], 1):
                cell = ws.cell(row=rr, column=ci, value=val)
                cell.font = Font(size=10, bold=(ci == 1))
                cell.alignment = Alignment(vertical="center", horizontal="left" if ci == 1 else "center")
                bd(cell)
                if rr % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            rr += 1

        for ci, val in enumerate(["NATION", total], 1):
            cell = ws.cell(row=rr, column=ci, value=val)
            cell.font = Font(size=10, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
            cell.alignment = Alignment(vertical="center", horizontal="left" if ci == 1 else "center")
            bd(cell)
        return rr + 2

    r = _write_table(r, "CLOSED BARRIERS BY TERRITORY (Closed This Month)" + live_tag, "closed_count")
    r = _write_table(r, "CLOSED BARRIERS BY TERRITORY (Closed This Quarter)" + live_tag, "closed_count_quarter")
    return r


def build_category_and_subcategory_tables(ws, r, all_td, taxonomy, live_tag=""):
    """
    Writes, starting at row r:
      1. BARRIER COUNTS BY TERRITORY — territory x category matrix,
         NATION row = nationwide per-category total.
      2. BARRIER SUBCATEGORY — TOTAL COUNTS — vertical list grouped under
         parent category via the taxonomy.
      3. SUBCATEGORY % OF CATEGORY — same layout, % of that subcategory's
         own parent category.
    Returns the row after the last table.
    """
    DARK_BLUE, MED_BLUE, MAROON = VIIV_NAVY_DARKEST, VIIV_NAVY_MED, VIIV_MAROON_DARK
    LIGHT_GREY, WHITE = "F2F2F2", "FFFFFF"

    def sh(cell, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    def bd(cell):
        s = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    cat1_counts = aggregate_category_counts(all_td, BARRIER_CAT_OUT_COLS["Barrier 1"])
    cat2_counts = aggregate_category_counts(all_td, BARRIER_CAT_OUT_COLS["Barrier 2"])
    subcat1_counts = aggregate_category_counts(all_td, BARRIER_SUBCAT_OUT_COLS["Barrier 1"])
    subcat2_counts = aggregate_category_counts(all_td, BARRIER_SUBCAT_OUT_COLS["Barrier 2"])

    # ---- Barrier Counts by Territory ----
    table_start_row = r
    all_categories = sorted(set(cat1_counts) | set(cat2_counts))
    n_cats = len(all_categories)

    terr_cat_counts = {}
    for label, *_ in TERRITORIES:
        td = all_td.get(label)
        if not td:
            terr_cat_counts[label] = {}
            continue
        b1 = tally_categories(td["campus_rows"], BARRIER_CAT_OUT_COLS["Barrier 1"])
        b2 = tally_categories(td["campus_rows"], BARRIER_CAT_OUT_COLS["Barrier 2"])
        terr_cat_counts[label] = dict(Counter(b1) + Counter(b2))

    if n_cats:
        end_letter = get_column_letter(1 + n_cats)
        ws.merge_cells(f"A{table_start_row}:{end_letter}{table_start_row}")
    title_cell = ws.cell(row=table_start_row, column=1)
    sh(title_cell, bg=DARK_BLUE, size=12)
    title_cell.value = "BARRIER COUNTS BY TERRITORY" + live_tag
    ws.row_dimensions[table_start_row].height = 24

    hdr_row = table_start_row + 1
    c0 = ws.cell(row=hdr_row, column=1, value="Territory")
    sh(c0, bg=MAROON, size=9)
    bd(c0)
    for ci, cat in enumerate(all_categories, 2):
        c = ws.cell(row=hdr_row, column=ci, value=cat)
        sh(c, bg=MAROON, size=9)
        bd(c)
    ws.row_dimensions[hdr_row].height = 40

    row = hdr_row + 1
    for label, *_ in TERRITORIES:
        counts = terr_cat_counts.get(label, {})
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(size=10, bold=True)
        label_cell.alignment = Alignment(vertical="center", horizontal="left")
        bd(label_cell)
        for ci, cat in enumerate(all_categories, 2):
            cell = ws.cell(row=row, column=ci, value=counts.get(cat, 0))
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            bd(cell)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
        ws.row_dimensions[row].height = 18
        row += 1

    nation_label_cell = ws.cell(row=row, column=1, value="NATION")
    nation_label_cell.font = Font(size=10, bold=True, color=WHITE)
    nation_label_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    nation_label_cell.alignment = Alignment(vertical="center", horizontal="left")
    bd(nation_label_cell)
    for ci, cat in enumerate(all_categories, 2):
        nation_total = sum(terr_cat_counts.get(label, {}).get(cat, 0) for label, *_ in TERRITORIES)
        cell = ws.cell(row=row, column=ci, value=nation_total)
        cell.font = Font(size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        bd(cell)
    ws.row_dimensions[row].height = 18
    r = row + 2

    # ---- Subcategory total counts, grouped by parent category ----
    ws.merge_cells(f"A{r}:D{r}")
    sh(ws[f"A{r}"], bg=DARK_BLUE, size=12)
    ws[f"A{r}"].value = "BARRIER SUBCATEGORY — TOTAL COUNTS" + live_tag
    ws.row_dimensions[r].height = 24
    r += 1

    for ci, h in enumerate(["Category", "Subcategory", "Barrier 1 Count", "Barrier 2 Count"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        sh(c, bg=MED_BLUE, size=9)
        bd(c)
    r += 1

    all_subcats = sorted(set(subcat1_counts) | set(subcat2_counts))
    grouped = {}
    for sub in all_subcats:
        cat = taxonomy.get(_normalize_header(sub), "Uncategorized")
        grouped.setdefault(cat, []).append(sub)

    for cat in sorted(grouped):
        subs = sorted(grouped[cat])
        cat_start_row = r
        for sub in subs:
            row_vals = [cat, sub, subcat1_counts.get(sub, 0), subcat2_counts.get(sub, 0)]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = Font(size=10, bold=(ci == 1))
                cell.alignment = Alignment(vertical="center", horizontal="left" if ci <= 2 else "center",
                                            wrap_text=(ci == 2))
                bd(cell)
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            r += 1
        if len(subs) > 1:
            ws.merge_cells(start_row=cat_start_row, start_column=1, end_row=r - 1, end_column=1)
            ws.cell(row=cat_start_row, column=1).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        r += 1

    # ---- Subcategory % of parent category ----
    ws.merge_cells(f"A{r}:C{r}")
    sh(ws[f"A{r}"], bg=DARK_BLUE, size=12)
    ws[f"A{r}"].value = "SUBCATEGORY % OF CATEGORY (Barrier 1 + Barrier 2 combined)" + live_tag
    ws.row_dimensions[r].height = 24
    r += 1

    for ci, h in enumerate(["Category", "Subcategory", "% of Category"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        sh(c, bg=MED_BLUE, size=9)
        bd(c)
    r += 1

    combined_subcat = Counter(subcat1_counts) + Counter(subcat2_counts)
    category_totals = Counter()
    for sub, cnt in combined_subcat.items():
        cat = taxonomy.get(_normalize_header(sub), "Uncategorized")
        category_totals[cat] += cnt

    for cat in sorted(grouped):
        cat_total = category_totals.get(cat, 0)
        subs = sorted(grouped[cat])
        cat_start_row = r
        for sub in subs:
            cnt = combined_subcat.get(sub, 0)
            pct = (cnt / cat_total) if cat_total else 0
            row_vals = [cat, sub, pct]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = Font(size=10, bold=(ci == 1))
                cell.alignment = Alignment(vertical="center", horizontal="left" if ci <= 2 else "center",
                                            wrap_text=(ci == 2))
                bd(cell)
                if ci == 3:
                    cell.number_format = "0%"
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            r += 1
        if len(subs) > 1:
            ws.merge_cells(start_row=cat_start_row, start_column=1, end_row=r - 1, end_column=1)
            ws.cell(row=cat_start_row, column=1).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        r += 1

    return r


def build_summary_sheet(wb, all_td, ws_totals, taxonomy):
    """
    Main entry point, called from main.py. Builds (replacing if it
    already exists) the "Trends & Takeaways" sheet from all_td (the
    per-territory data dict) and ws_totals (the VCART Totals worksheet,
    for reading its time series).
    """
    print("  Building Trends & Takeaways sheet...")

    if "Trends & Takeaways" in wb.sheetnames:
        del wb["Trends & Takeaways"]
    ws = wb.create_sheet("Trends & Takeaways")

    valid_td = [td for td in all_td.values() if td]
    total_campuses = sum(td["campus_count"] for td in valid_td)

    DARK_BLUE, MED_BLUE, MAROON, MAROON_MED = VIIV_NAVY_DARKEST, VIIV_NAVY_MED, VIIV_MAROON_DARK, VIIV_MAROON_MED
    WHITE, LIGHT_GREY = "FFFFFF", "F2F2F2"

    def sh(cell, bg=DARK_BLUE, fg=WHITE, size=10, bold=True):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    def bd(cell):
        s = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

    # Same numeric/bool guard as compute_total_barriers() in main.py — kept
    # in sync manually since this needs the per-barrier-type breakdown
    # (top_bi/top_pct below), not just the final sum.
    barrier_totals = [
        sum(td["totals"].get(BARRIER_COLS_OUT[bi], 0) or 0 for td in valid_td
            if isinstance(td["totals"].get(BARRIER_COLS_OUT[bi]), (int, float))
            and not isinstance(td["totals"].get(BARRIER_COLS_OUT[bi]), bool))
        for bi in range(len(BARRIER_COLS_OUT))
    ]
    top_bi = barrier_totals.index(max(barrier_totals)) if barrier_totals else 0
    top_pct = barrier_totals[top_bi] / total_campuses if total_campuses else 0
    total_barriers = sum(barrier_totals)

    pathway_total = sum(td["totals"].get(29, 0) or 0 for td in valid_td if isinstance(td["totals"].get(29), (int, float)))
    pathway_pct = pathway_total / total_campuses if total_campuses else 0

    # Days to Close data currently exists (scanned per-campus at read
    # time), but months_data below only pulls what's in the time series
    # (barrier %s), matching what update_time_series actually writes.
    #
    # FIX: Months/Total Barriers now live at TS_MONTH_COL/TS_BARRIERS_COL
    # (D/E) instead of cols 1-2 — reading the old position here silently
    # returned nothing once main.py's TS table moved, showing a
    # "0-Month Trend" with no rows even though the source data was fine.
    months_data = []
    for src_r in range(TS_START_ROW, TS_END_ROW + 1):
        month_val = ws_totals.cell(row=src_r, column=TS_MONTH_COL).value
        if not month_val:
            continue
        if not any(m.lower() in str(month_val).lower() for m in MONTH_NAMES):
            continue
        total_b = ws_totals.cell(row=src_r, column=TS_BARRIERS_COL).value
        barrier_pcts = [ws_totals.cell(row=src_r, column=bc).value
                        for bc in BARRIER_COLS_OUT]
        months_data.append((str(month_val), total_b, barrier_pcts))
    months_written = len(months_data)

    pull_date_str = datetime.now().strftime("%m/%d/%y")
    LIVE_TAG = f"  (LIVE DATA — as of {pull_date_str})"
    HIST_TAG = f"  (HISTORICAL — {months_written}-Month Trend)"

    ws.merge_cells("A1:T1")
    ws["A1"].value = "VCART Trends & Takeaways"
    ws["A1"].font = Font(bold=True, size=22, color=DARK_BLUE)
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:T2")
    ws["A2"].value = (f"Refreshed: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
                       f"    |    Total Campuses: {total_campuses}"
                       f"    |    Current Total Barriers: {total_barriers}")
    ws["A2"].font = Font(italic=True, size=11, color="646464")
    ws["A2"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[2].height = 22

    r = 4
    ws.merge_cells(f"A{r}:T{r}")
    sh(ws[f"A{r}"], bg=DARK_BLUE, size=12)
    ws[f"A{r}"].value = f"KEY METRICS  —  {datetime.now().strftime('%B %d, %Y  %I:%M %p')}{LIVE_TAG}"
    ws.row_dimensions[r].height = 24
    r += 1

    metrics = [
        ("Total Barriers",   str(total_barriers),                        MED_BLUE),
        ("Top Barrier Type", f"{BARRIER_NAMES[top_bi]} ({top_pct:.0%})", MAROON_MED),
        ("Pathway Mapped",   f"{pathway_pct:.0%} of campuses",           MAROON),
    ]
    card_cols = [(1, 6), (7, 12), (13, 18)]
    for (start_c, end_c), (label, value, color) in zip(card_cols, metrics):
        ws.merge_cells(start_row=r, start_column=start_c, end_row=r, end_column=end_c)
        cell = ws.cell(row=r, column=start_c, value=label)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r + 1, start_column=start_c, end_row=r + 1, end_column=end_c)
        cell2 = ws.cell(row=r + 1, column=start_c, value=value)
        cell2.font = Font(bold=True, size=13, color=color)
        cell2.fill = PatternFill("solid", fgColor="F2F2F2")
        cell2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 18
    ws.row_dimensions[r + 1].height = 28
    r += 3

    # ---- Barrier Trends — Monthly ----
    ws.merge_cells(f"A{r}:I{r}")
    sh(ws[f"A{r}"], bg=DARK_BLUE, size=12)
    ws[f"A{r}"].value = "BARRIER TRENDS — MONTHLY" + HIST_TAG
    ws.row_dimensions[r].height = 24
    r += 1

    trend_hdrs = (["Month", "Total\nBarriers", "MoM\nChange"] +
                  BARRIER_NAMES_WRAPPED)
    for ci, h in enumerate(trend_hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        sh(c, bg=MED_BLUE, size=9)
        bd(c)
    ws.row_dimensions[r].height = 40
    trend_hdr_row = r
    r += 1

    prev_total = None
    for month_name, total_b, barrier_pcts in months_data:
        mom = None
        if isinstance(prev_total, (int, float)) and isinstance(total_b, (int, float)):
            mom = total_b - prev_total
        row_vals = [month_name, total_b, mom] + barrier_pcts
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            bd(cell)
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            if ci == 3 and val is not None:
                cell.font = Font(size=10, bold=True,
                                  color="375623" if val < 0 else ("C00000" if val > 0 else "000000"))
            # FIX: these are raw monthly counts (same source as Total
            # Barriers — sums written by update_time_series), not
            # fractions. "0%" was displaying e.g. 38 as "3800%". The
            # variable name barrier_pcts is a holdover from an earlier
            # design; "0" is the correct format for what's actually here.
            if ci > 3 and val is not None:
                cell.number_format = "0"
        ws.row_dimensions[r].height = 16
        prev_total = total_b if isinstance(total_b, (int, float)) else prev_total
        r += 1
    trend_data_end = r - 1
    r += 1

    if months_written >= 2:
        chart1 = LineChart()
        chart1.title = "Total Barriers Month-over-Month" + HIST_TAG
        chart1.style = 2
        chart1.height = 18
        chart1.width = 28
        chart1.y_axis.title = "Total Barriers"
        chart1.y_axis.numFmt = "0"
        chart1.x_axis.title = "Month"
        data1 = Reference(ws, min_col=2, max_col=2, min_row=trend_hdr_row, max_row=trend_data_end)
        chart1.add_data(data1, titles_from_data=True)
        chart1.series[0].graphicalProperties.line.solidFill = "1F497D"
        chart1.series[0].graphicalProperties.line.width = 28000
        chart1.series[0].marker.symbol = "circle"
        chart1.series[0].marker.size = 7
        chart1.set_categories(Reference(ws, min_col=1, min_row=trend_hdr_row + 1, max_row=trend_data_end))
        ws.add_chart(chart1, "V4")

    # ---- Barrier Breakdown by Territory (% of Campuses) ----
    ws.merge_cells(f"A{r}:H{r}")
    sh(ws[f"A{r}"], bg=DARK_BLUE, size=12)
    ws[f"A{r}"].value = "BARRIER BREAKDOWN BY TERRITORY (% of Campuses)" + LIVE_TAG
    ws.row_dimensions[r].height = 24
    r += 1

    breakdown_hdrs = ["Territory", "Total\nCampuses"] + BARRIER_NAMES_WRAPPED
    for ci, h in enumerate(breakdown_hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        sh(c, bg=MED_BLUE if ci > 1 else DARK_BLUE, size=9)
        bd(c)
    ws.row_dimensions[r].height = 40
    r += 1

    for label, *_ in TERRITORIES:
        td = all_td.get(label)
        tot = td["campus_count"] if td else 0
        barrier_pcts = [((td["totals"].get(bc, 0) or 0) / tot if tot and td else 0)
                         for bc in BARRIER_COLS_OUT]
        row_vals = [label, tot] + barrier_pcts
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = Font(size=10, bold=(ci == 1))
            cell.alignment = Alignment(vertical="center", horizontal="center")
            bd(cell)
            if ci > 2 and isinstance(val, float):
                cell.number_format = "0%"
                cell.fill = PatternFill("solid", fgColor=(
                    "FFC7CE" if val > 0.6 else ("FFF2CC" if val > 0.3 else "E2EFDA")))
                cell.font = Font(size=10, bold=False, color=(
                    "9C0006" if val > 0.6 else ("7D6608" if val > 0.3 else "375623")))
        ws.row_dimensions[r].height = 18
        r += 1

    nation_pcts = [sum(td["totals"].get(bc, 0) or 0 for td in valid_td) / total_campuses if total_campuses else 0
                   for bc in BARRIER_COLS_OUT]
    nation_vals = ["NATION", total_campuses] + nation_pcts
    for ci, val in enumerate(nation_vals, 1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font = Font(size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        bd(cell)
        if ci > 2 and isinstance(val, float):
            cell.number_format = "0%"
    ws.row_dimensions[r].height = 18
    r += 2

    # ---- Category / subcategory tables ----
    r = build_category_and_subcategory_tables(ws, r, all_td, taxonomy, live_tag=LIVE_TAG)

    # ---- Closed Barriers by Territory ----
    r = build_closed_barriers_table(ws, r, all_td, live_tag=LIVE_TAG)

    # ---- Days to Close by Category ----
    r = build_days_to_close_table(ws, r, all_td, live_tag=LIVE_TAG)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    # 14, not the old 11 — "Reimbursement" (13 chars, one line of the
    # BARRIER_NAMES_WRAPPED header above) needs the room, or Excel wraps
    # it again mid-word on top of the forced line break.
    for col in ["D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 14
    for col in ["J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        ws.column_dimensions[col].width = 9

    ws.sheet_view.showGridLines = False

    print(f"    Trends & Takeaways built — {months_written} months of data, "
          f"category/subcategory + closed-barrier tables added")